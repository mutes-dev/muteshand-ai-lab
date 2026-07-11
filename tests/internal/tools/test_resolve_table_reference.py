"""Tests for resolve_table_reference tool."""

import csv
import os
import sys
import tempfile

import pytest

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "..", ".."))

from tools.resolve_table_reference import run as resolve_table_reference


def _project_root():
    return os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", ".."))


def _project_tmp():
    return os.path.join(_project_root(), "tmp")


def _make_csv(rows, headers=None):
    fd, path = tempfile.mkstemp(dir=_project_tmp(), suffix=".csv")
    with os.fdopen(fd, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        if headers:
            writer.writerow(headers)
        writer.writerows(rows)
    return path


def _make_xlsx(rows, headers=None, sheet_name="Sheet1"):
    from openpyxl import Workbook

    fd, path = tempfile.mkstemp(dir=_project_tmp(), suffix=".xlsx")
    os.close(fd)
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    if headers:
        ws.append(headers)
    for row in rows:
        ws.append(row)
    wb.save(path)
    wb.close()
    return path


def _rel(path):
    return os.path.relpath(path, _project_root())


class TestResolveTableReference:
    def test_csv_row_resolution(self):
        headers = ["Name", "Age", "City"]
        rows = [["Alice", "30", "NYC"], ["Bob", "25", "LA"], ["Carol", "35", "Chicago"]]
        path = _make_csv(rows, headers=headers)
        try:
            result = resolve_table_reference(_rel(path), reference_type="row", row_number=2)
            assert result["status"] == "success"
            assert result["reference_type"] == "row"
            assert result["row_number"] == 2
            assert result["row"] == ["Alice", "30", "NYC"]
            assert result["header_map"] == {"Name": 1, "Age": 2, "City": 3}
            assert result["data_ref"]["row_number"] == 2
        finally:
            os.unlink(path)

    def test_xlsx_row_resolution(self):
        headers = ["Product", "Price", "Quantity"]
        rows = [["Apple", "1.50", "10"], ["Banana", "0.75", "20"], ["Cherry", "2.00", "15"]]
        path = _make_xlsx(rows, headers=headers, sheet_name="Sales")
        try:
            result = resolve_table_reference(
                _rel(path), reference_type="row", row_number=2, sheet_name="Sales"
            )
            assert result["status"] == "success"
            assert result["file_type"] == "xlsx"
            assert result["sheet_name"] == "Sales"
            assert result["row"] == ["Apple", "1.50", "10"]
        finally:
            os.unlink(path)

    def test_xlsx_cell_address_resolution(self):
        headers = ["A", "B", "C"]
        rows = [["1", "2", "3"], ["4", "5", "6"], ["7", "8", "9"]]
        path = _make_xlsx(rows, headers=headers, sheet_name="Data")
        try:
            result = resolve_table_reference(
                _rel(path), reference_type="cell", cell_address="B2", sheet_name="Data"
            )
            assert result["status"] == "success"
            assert result["value"] == "2"
            assert result["cell_address"] == "B2"
            assert result["column_index"] == 2
            assert result["row_number"] == 2
        finally:
            os.unlink(path)

    def test_csv_column_name_cell_resolution(self):
        headers = ["Name", "Age", "City"]
        rows = [["Alice", "30", "NYC"], ["Bob", "25", "LA"], ["Carol", "35", "Chicago"]]
        path = _make_csv(rows, headers=headers)
        try:
            result = resolve_table_reference(
                _rel(path), reference_type="cell", row_number=3, column_name="Age"
            )
            assert result["status"] == "success"
            assert result["value"] == "25"
            assert result["column_name"] == "Age"
            assert result["column_index"] == 2
        finally:
            os.unlink(path)

    def test_entity_from_row(self):
        headers = ["Name", "Age", "City"]
        rows = [["Alice", "30", "NYC"], ["Bob", "25", "LA"], ["Carol", "35", "Chicago"]]
        path = _make_csv(rows, headers=headers)
        try:
            result = resolve_table_reference(
                _rel(path), reference_type="entity_from_row", row_number=3, entity_column="City"
            )
            assert result["status"] == "success"
            assert result["entity"] == "LA"
            assert result["value"] == "LA"
            assert result["row_number"] == 3
            assert result["data_ref"]["column_name"] == "City"
        finally:
            os.unlink(path)

    def test_missing_row_error(self):
        headers = ["Name", "Age"]
        rows = [["Alice", "30"], ["Bob", "25"]]
        path = _make_csv(rows, headers=headers)
        try:
            result = resolve_table_reference(_rel(path), reference_type="row", row_number=100)
            assert result["status"] == "error"
            assert result["error_code"] == "row_not_found"
        finally:
            os.unlink(path)

    def test_missing_column_error(self):
        headers = ["Name", "Age"]
        rows = [["Alice", "30"], ["Bob", "25"]]
        path = _make_csv(rows, headers=headers)
        try:
            result = resolve_table_reference(
                _rel(path), reference_type="cell", row_number=2, column_name="Salary"
            )
            assert result["status"] == "error"
            assert result["error_code"] == "missing_column"
        finally:
            os.unlink(path)

    def test_missing_sheet_error(self):
        headers = ["A"]
        rows = [["1"]]
        path = _make_xlsx(rows, headers=headers, sheet_name="Sheet1")
        try:
            result = resolve_table_reference(
                _rel(path), reference_type="row", row_number=2, sheet_name="MissingSheet"
            )
            assert result["status"] == "error"
            assert result["error_code"] == "missing_sheet"
        finally:
            os.unlink(path)

    def test_unsupported_extension_error(self):
        fd, path = tempfile.mkstemp(dir=_project_tmp(), suffix=".txt")
        with os.fdopen(fd, "w", encoding="utf-8") as f:
            f.write("not a table\n")
        try:
            result = resolve_table_reference(
                _rel(path), reference_type="row", row_number=1
            )
            assert result["status"] == "error"
            assert result["error_code"] == "unsupported_format"
        finally:
            os.unlink(path)

    def test_formula_cell_does_not_execute_formula(self):
        from openpyxl import Workbook

        fd, path = tempfile.mkstemp(dir=_project_tmp(), suffix=".xlsx")
        os.close(fd)
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Value"
        ws["A2"] = 5
        ws["A3"] = "=A2+10"
        wb.save(path)
        wb.close()
        try:
            result = resolve_table_reference(
                _rel(path), reference_type="cell", cell_address="A3"
            )
            assert result["status"] == "success"
            # If the formula was executed, value would be 15; it must not be executed.
            assert result["value"] != "15"
            # It should be either empty or the cached formula string.
            assert "Formula" in result["value"] or result["value"] == "" or result["value"] is None
        finally:
            os.unlink(path)

    def test_bounds_applied(self):
        headers = ["Name", "Age"]
        rows = [[f"Person{i}", str(i)] for i in range(1, 2500)]
        path = _make_csv(rows, headers=headers)
        try:
            result = resolve_table_reference(
                _rel(path),
                reference_type="row",
                row_number=1500,
                max_rows_scan=100,
                max_columns_scan=1,
                max_cell_chars=5,
            )
            assert result["status"] == "error"
            assert result["error_code"] == "row_not_found"
            assert result["bounds_applied"]["max_rows_scan"] == 100
        finally:
            os.unlink(path)

    def test_no_analysis_output(self):
        headers = ["Name", "Score"]
        rows = [["Alice", "90"], ["Bob", "80"], ["Carol", "100"]]
        path = _make_csv(rows, headers=headers)
        try:
            result = resolve_table_reference(
                _rel(path), reference_type="cell", row_number=3, column_name="Score"
            )
            assert result["status"] == "success"
            assert "sum" not in result
            assert "average" not in result
            assert "max" not in result
            assert "min" not in result
            assert "highest" not in result
            assert "filter" not in result
            assert "sort" not in result
            assert result["value"] == "80"
        finally:
            os.unlink(path)


    def test_case_insensitive_column_name_lookup(self):
        headers = ["Name", "Age", "City"]
        rows = [["Alice", "30", "NYC"], ["Bob", "25", "LA"], ["Carol", "35", "Chicago"]]
        path = _make_csv(rows, headers=headers)
        try:
            result = resolve_table_reference(
                _rel(path), reference_type="entity_from_row", row_number=3, entity_column="name"
            )
            assert result["status"] == "success"
            assert result["entity"] == "Bob"
            assert result["column_name"] == "Name"
        finally:
            os.unlink(path)

    def test_auto_name_like_column_unique(self):
        headers = ["Name", "Age", "Score"]
        rows = [["Alice", "30", "90"], ["Bob", "25", "80"], ["Cara", "35", "97"]]
        path = _make_csv(rows, headers=headers)
        try:
            result = resolve_table_reference(
                _rel(path),
                reference_type="entity_from_row",
                row_number=4,
                entity_column="__AUTO_NAME_LIKE__",
            )
            assert result["status"] == "success"
            assert result["entity"] == "Cara"
            assert result["column_name"] == "Name"
        finally:
            os.unlink(path)

    def test_auto_name_like_column_ambiguous(self):
        headers = ["Name", "FullName", "Score"]
        rows = [["Alice", "A. Smith", "90"], ["Bob", "B. Jones", "80"]]
        path = _make_csv(rows, headers=headers)
        try:
            result = resolve_table_reference(
                _rel(path),
                reference_type="entity_from_row",
                row_number=3,
                entity_column="__AUTO_NAME_LIKE__",
            )
            assert result["status"] == "error"
            assert result["error_code"] == "ambiguous_name_like_column"
        finally:
            os.unlink(path)

    def test_auto_name_like_column_absent(self):
        headers = ["Age", "Score"]
        rows = [["30", "90"], ["25", "80"]]
        path = _make_csv(rows, headers=headers)
        try:
            result = resolve_table_reference(
                _rel(path),
                reference_type="entity_from_row",
                row_number=3,
                entity_column="__AUTO_NAME_LIKE__",
            )
            assert result["status"] == "error"
            assert result["error_code"] == "ambiguous_name_like_column"
        finally:
            os.unlink(path)

    def test_xlsx_auto_name_like_column_unique(self):
        headers = ["Name", "Score"]
        rows = [["Alice", "90"], ["Bob", "80"], ["Cara", "97"]]
        path = _make_xlsx(rows, headers=headers, sheet_name="People")
        try:
            result = resolve_table_reference(
                _rel(path),
                reference_type="entity_from_row",
                row_number=4,
                entity_column="__AUTO_NAME_LIKE__",
                sheet_name="People",
            )
            assert result["status"] == "success"
            assert result["entity"] == "Cara"
        finally:
            os.unlink(path)

    # === F2B-2-FIX1: header-row cell/row resolution ===

    def test_csv_cell_c1_header_row(self):
        headers = ["Name", "Score", "Team"]
        rows = [["Alice", "91", "Alpha"], ["Bob", "82", "Bravo"], ["Cara", "97", "Gamma"]]
        path = _make_csv(rows, headers=headers)
        try:
            result = resolve_table_reference(
                _rel(path), reference_type="cell", cell_address="C1"
            )
            assert result["status"] == "success"
            assert result["value"] == "Team"
            assert result["column_name"] == "Team"
        finally:
            os.unlink(path)

    def test_csv_cell_b1_header_row(self):
        headers = ["Name", "Score", "Team"]
        rows = [["Alice", "91", "Alpha"], ["Bob", "82", "Bravo"]]
        path = _make_csv(rows, headers=headers)
        try:
            result = resolve_table_reference(
                _rel(path), reference_type="cell", cell_address="B1"
            )
            assert result["status"] == "success"
            assert result["value"] == "Score"
        finally:
            os.unlink(path)

    def test_csv_cell_a1_header_row(self):
        headers = ["Name", "Score", "Team"]
        rows = [["Alice", "91", "Alpha"], ["Bob", "82", "Bravo"]]
        path = _make_csv(rows, headers=headers)
        try:
            result = resolve_table_reference(
                _rel(path), reference_type="cell", cell_address="A1"
            )
            assert result["status"] == "success"
            assert result["value"] == "Name"
        finally:
            os.unlink(path)

    def test_csv_row_one_header_row(self):
        headers = ["Name", "Score", "Team"]
        rows = [["Alice", "91", "Alpha"], ["Bob", "82", "Bravo"]]
        path = _make_csv(rows, headers=headers)
        try:
            result = resolve_table_reference(
                _rel(path), reference_type="row", row_number=1
            )
            assert result["status"] == "success"
            assert result["row"] == ["Name", "Score", "Team"]
        finally:
            os.unlink(path)

    def test_xlsx_cell_c1_header_row(self):
        headers = ["Name", "Score", "Team"]
        rows = [["Alice", "91", "Alpha"], ["Bob", "82", "Bravo"], ["Cara", "97", "Gamma"]]
        path = _make_xlsx(rows, headers=headers, sheet_name="People")
        try:
            result = resolve_table_reference(
                _rel(path), reference_type="cell", cell_address="C1", sheet_name="People"
            )
            assert result["status"] == "success"
            assert result["value"] == "Team"
        finally:
            os.unlink(path)

    def test_system_entry_cell_c1_header_row(self):
        headers = ["Name", "Score", "Team"]
        rows = [["Alice", "91", "Alpha"], ["Bob", "82", "Bravo"], ["Cara", "97", "Gamma"]]
        path = _make_csv(rows, headers=headers)
        try:
            rel_path = _rel(path)
            tool_call = f'resolve_table_reference \"{rel_path}\" \"cell\" \"\" \"1\" 1 0 \"C1\" \"\" 0 \"\" 0 0 0'
            from system.entry.system_entry import system_entry
            result = system_entry(tool_call)
            assert result["status"] == "success"
            assert result["result"] is not None
            assert result["result"].get("value") == "Team"
        finally:
            os.unlink(path)


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
