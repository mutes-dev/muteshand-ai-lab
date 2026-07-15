"""Tests for tools/analyze_table.py — F5A bounded deterministic table analysis."""

import csv
import os
import sys
import tempfile

import pytest

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "..", ".."))

from tools import analyze_table


# Test the raw payload implementation directly (not the system_entry wrapper).
def _call(*args, **kwargs):
    return analyze_table._analyze_table_impl(*args, **kwargs)


BASE_PATH = os.path.abspath("E:/MutesHand")


@pytest.fixture
def table_tmp_path():
    """Create a temporary directory inside the project root for safe file tests."""
    with tempfile.TemporaryDirectory(dir=BASE_PATH) as d:
        yield d


def _write_csv(file_path: str, rows: list[list]):
    with open(file_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        for row in rows:
            writer.writerow(row)


def _rel(path: str) -> str:
    return os.path.relpath(path, BASE_PATH)


class TestAnalyzeTableCsvSuccess:
    def test_count_rows(self, table_tmp_path):
        p = os.path.join(table_tmp_path, "scores.csv")
        _write_csv(p, [["Name", "Score"], ["Alice", "10"], ["Bob", "20"], ["", ""]])
        result = _call(_rel(p), "count_rows")
        assert result["status"] == "success"
        assert result["operation"] == "count_rows"
        assert result["computed_value"] == 2
        assert result["rows_evaluated"] == 3
        assert result["column_refs"] == []

    def test_max_integer(self, table_tmp_path):
        p = os.path.join(table_tmp_path, "scores.csv")
        _write_csv(p, [["Name", "Score"], ["Alice", "85"], ["Bob", "91"], ["Cara", "78"]])
        result = _call(_rel(p), "max", "Score")
        assert result["status"] == "success"
        assert result["computed_value"] == "91"
        assert result["value_kind"] == "integer"
        assert result["extreme_cells"] == ["B3"]

    def test_min_with_explicit_associated_column(self, table_tmp_path):
        p = os.path.join(table_tmp_path, "scores.csv")
        _write_csv(p, [["Name", "Score"], ["Alice", "85"], ["Bob", "91"], ["Cara", "78"]])
        result = _call(_rel(p), "min", "Score", associated_column="Name")
        assert result["status"] == "success"
        assert result["computed_value"] == "78"
        assert result["associated"]["associated_column"] == "Name"
        assert result["answer_text"].endswith("Cara.")

    def test_auto_name_like_associated_column(self, table_tmp_path):
        p = os.path.join(table_tmp_path, "people.csv")
        _write_csv(p, [["Full Name", "Score"], ["Bob", "50"], ["Alice", "90"]])
        result = _call(_rel(p), "max", "Score", associated_column="__AUTO_NAME_LIKE__")
        assert result["status"] == "success"
        assert result["associated"]["associated_column"] == "Full Name"
        assert "Alice" in result["answer_text"]

    def test_sum_decimal(self, table_tmp_path):
        p = os.path.join(table_tmp_path, "amounts.csv")
        _write_csv(p, [["Amount"], ["10.5"], ["2.25"], ["3"]])
        result = _call(_rel(p), "sum", "Amount")
        assert result["status"] == "success"
        assert result["computed_value"] == "15.75"
        assert result["value_kind"] == "decimal"

    def test_average_terminating_exact(self, table_tmp_path):
        p = os.path.join(table_tmp_path, "vals.csv")
        _write_csv(p, [["Val"], ["10"], ["20"], ["30"]])
        result = _call(_rel(p), "average", "Val")
        assert result["status"] == "success"
        assert result["computed_value"] == "20"
        assert result["value_kind"] == "integer"
        assert result["rounded"] is False
        assert result["computed_sum"] == "60"
        assert result["computed_count"] == 3
        assert result["rounding_mode"] == "ROUND_HALF_EVEN"

    def test_average_repeating_decimal_is_rounded(self, table_tmp_path):
        p = os.path.join(table_tmp_path, "vals.csv")
        _write_csv(p, [["Val"], ["10"], ["20"], ["31"]])
        result = _call(_rel(p), "average", "Val")
        assert result["status"] == "success"
        assert result["value_kind"] == "decimal"
        assert result["rounded"] is True
        assert result["precision"] == 12
        # 61/3 = 20.3333333333...; 12 significant digits => "20.3333333333".
        assert result["computed_value"].startswith("20.3333333333")
        assert "approximately" in result["answer_text"]
        assert result["computed_sum"] == "61"
        assert result["computed_count"] == 3

    def test_tied_rows_reported(self, table_tmp_path):
        p = os.path.join(table_tmp_path, "tied.csv")
        _write_csv(p, [["Name", "Score"], ["A", "10"], ["B", "10"], ["C", "5"]])
        result = _call(_rel(p), "max", "Score", associated_column="Name")
        assert result["status"] == "success"
        assert result["tied_row_count"] == 2
        assert sorted(result["tied_rows"]) == [2, 3]
        assert "A, B" in result["answer_text"] or "B, A" in result["answer_text"]

    def test_blank_cells_excluded(self, table_tmp_path):
        p = os.path.join(table_tmp_path, "blanks.csv")
        _write_csv(p, [["Score"], ["10"], [""], ["20"]])
        result = _call(_rel(p), "sum", "Score")
        assert result["status"] == "success"
        assert result["computed_value"] == "30"
        assert result["blank_cells"] == 1
        assert result["numeric_cells"] == 2


class TestAnalyzeTableXlsxSuccess:
    def test_max_xlsx(self, table_tmp_path):
        pytest.importorskip("openpyxl")
        from openpyxl import Workbook

        p = os.path.join(table_tmp_path, "scores.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(["Name", "Score"])
        ws.append(["Alice", 85])
        ws.append(["Bob", 91])
        wb.save(p)

        result = _call(_rel(p), "max", "Score")
        assert result["status"] == "success"
        assert result["computed_value"] == "91"
        assert result["sheet_name"] == "Sheet1"

    def test_sheet_name_extraction(self, table_tmp_path):
        pytest.importorskip("openpyxl")
        from openpyxl import Workbook

        p = os.path.join(table_tmp_path, "multi.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "Grades"
        ws.append(["Score"])
        ws.append([95])
        wb.create_sheet("Other")
        wb.save(p)

        result = _call(_rel(p), "max", "Score", sheet_name="Grades")
        assert result["status"] == "success"
        assert result["computed_value"] == "95"
        assert result["sheet_name"] == "Grades"

    def test_formula_cell_is_blocked(self, table_tmp_path):
        pytest.importorskip("openpyxl")
        from openpyxl import Workbook

        p = os.path.join(table_tmp_path, "formula.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.append(["Score"])
        ws.append(["=SUM(1,2)"])
        wb.save(p)

        result = _call(_rel(p), "sum", "Score")
        assert result["status"] == "unsupported"
        assert result["status_reason"] == "formula_cell_present"


class TestAnalyzeTableErrors:
    def test_missing_file(self, table_tmp_path):
        p = os.path.join(table_tmp_path, "missing.csv")
        result = _call(_rel(p), "count_rows")
        assert result["status"] == "not_found"
        assert result["status_reason"] == "file_not_found"

    def test_unsupported_file_type(self, table_tmp_path):
        p = os.path.join(table_tmp_path, "legacy.xls")
        with open(p, "w") as f:
            f.write("dummy")
        result = _call(_rel(p), "count_rows")
        assert result["status"] == "unsupported"
        assert result["status_reason"] == "unsupported_file_type"

    def test_missing_column(self, table_tmp_path):
        p = os.path.join(table_tmp_path, "scores.csv")
        _write_csv(p, [["Name", "Score"], ["A", "10"]])
        result = _call(_rel(p), "max", "Grade")
        assert result["status"] == "not_found"
        assert result["status_reason"] == "column_not_found"

    def test_duplicate_column_header(self, table_tmp_path):
        p = os.path.join(table_tmp_path, "dup.csv")
        _write_csv(p, [["Score", "Score"], ["A", "10"]])
        result = _call(_rel(p), "max", "Score")
        assert result["status"] == "ambiguous"
        assert result["status_reason"] == "duplicate_column_header"

    def test_non_numeric_value(self, table_tmp_path):
        p = os.path.join(table_tmp_path, "bad.csv")
        _write_csv(p, [["Score"], ["ten"]])
        result = _call(_rel(p), "max", "Score")
        assert result["status"] == "unsupported"
        assert result["status_reason"] == "non_numeric_value_present"

    def test_no_numeric_values(self, table_tmp_path):
        p = os.path.join(table_tmp_path, "empty.csv")
        _write_csv(p, [["Score"], [""]])
        result = _call(_rel(p), "max", "Score")
        assert result["status"] == "not_found"
        assert result["status_reason"] == "no_numeric_values"

    def test_multiple_sheets_requires_selection(self, table_tmp_path):
        pytest.importorskip("openpyxl")
        from openpyxl import Workbook

        p = os.path.join(table_tmp_path, "multi.xlsx")
        wb = Workbook()
        wb.active.title = "A"
        wb.create_sheet("B")
        wb.save(p)

        result = _call(_rel(p), "count_rows")
        assert result["status"] == "ambiguous"
        assert result["status_reason"] == "multiple_sheets_require_selection"

    def test_missing_sheet(self, table_tmp_path):
        pytest.importorskip("openpyxl")
        from openpyxl import Workbook

        p = os.path.join(table_tmp_path, "book.xlsx")
        wb = Workbook()
        wb.active.title = "Sheet1"
        wb.save(p)

        result = _call(_rel(p), "count_rows", sheet_name="Missing")
        assert result["status"] == "not_found"
        assert result["status_reason"] == "sheet_not_found"

    def test_analysis_bounds_exceeded_rows(self, table_tmp_path, monkeypatch):
        p = os.path.join(table_tmp_path, "big.csv")
        _write_csv(p, [["Score"], ["1"], ["2"], ["3"]])
        monkeypatch.setattr(analyze_table, "MAX_DATA_ROWS", 1)
        result = _call(_rel(p), "sum", "Score")
        assert result["status"] == "unsupported"
        assert result["status_reason"] == "analysis_bounds_exceeded"

    def test_analysis_bounds_exceeded_columns(self, table_tmp_path, monkeypatch):
        p = os.path.join(table_tmp_path, "wide.csv")
        headers = [f"Col{i}" for i in range(101)]
        _write_csv(p, [headers, ["1"] * 101])
        result = _call(_rel(p), "sum", "Col0")
        assert result["status"] == "unsupported"
        assert result["status_reason"] == "analysis_bounds_exceeded"

    def test_path_traversal_blocked(self, table_tmp_path):
        result = _call("../outside.csv", "count_rows")
        assert result["status"] == "blocked"
        assert result["status_reason"] == "path_safety_blocked"

    def test_unsupported_operation(self, table_tmp_path):
        p = os.path.join(table_tmp_path, "scores.csv")
        _write_csv(p, [["Score"], ["10"]])
        result = _call(_rel(p), "median", "Score")
        assert result["status"] == "unsupported"
        assert result["status_reason"] == "unsupported_operation"

    def test_missing_target_column(self, table_tmp_path):
        p = os.path.join(table_tmp_path, "scores.csv")
        _write_csv(p, [["Score"], ["10"]])
        result = _call(_rel(p), "max")
        assert result["status"] == "unsupported"
        assert result["status_reason"] == "missing_target_column"

    def test_auto_name_like_ambiguous(self, table_tmp_path):
        p = os.path.join(table_tmp_path, "ambig.csv")
        _write_csv(p, [["First Name", "Last Name", "Score"], ["A", "B", "10"]])
        result = _call(_rel(p), "max", "Score", associated_column="__AUTO_NAME_LIKE__")
        assert result["status"] == "ambiguous"
        assert result["status_reason"] == "associated_column_ambiguous"

class TestAnalyzeTableOverview:
    def test_overview_returns_table_and_numeric_column_stats(self, table_tmp_path):
        p = os.path.join(table_tmp_path, "overview.csv")
        _write_csv(p, [["Name", "Score"], ["Alice", "84"], ["Bob", "91"], ["Cara", "87"]])
        result = _call(_rel(p), "overview")
        assert result["status"] == "success"
        assert result["operation"] == "overview"
        assert result["data_row_count"] == 3
        assert result["blank_row_count"] == 0
        assert result["column_count"] == 2
        assert result["column_names"] == ["Name", "Score"]
        assert result["answer_text"].startswith("The table contains 3 data rows and 2 columns.")
        assert "See Details / Evidence" in result["answer_text"]

        score_col = next(c for c in result["columns"] if c["column_name"] == "Score")
        assert score_col["classification"] == "numeric"
        assert score_col["numeric_count"] == 3
        assert score_col["min"] == "84"
        assert score_col["max"] == "91"
        assert score_col["sum"] == "262"
        assert score_col["average"] == "87.3333333333"
        assert score_col["average_rounded"] is True
        assert "B2:B4" in score_col["contributing_range"]

    def test_overview_text_column_counts(self, table_tmp_path):
        p = os.path.join(table_tmp_path, "text_col.csv")
        _write_csv(p, [["Name", "Score"], ["Alice", "10"], ["", "20"], ["Bob", ""]])
        result = _call(_rel(p), "overview")
        assert result["status"] == "success"
        name_col = next(c for c in result["columns"] if c["column_name"] == "Name")
        assert name_col["classification"] == "text"
        assert name_col["nonblank_count"] == 2
        assert name_col["blank_count"] == 1
        assert name_col["distinct_count"] == 2

    def test_overview_mixed_column_is_warned(self, table_tmp_path):
        p = os.path.join(table_tmp_path, "mixed.csv")
        _write_csv(p, [["Val"], ["10"], ["bad"], ["20"]])
        result = _call(_rel(p), "overview")
        assert result["status"] == "success"
        col = result["columns"][0]
        assert col["classification"] == "mixed"
        assert "numeric statistics omitted" in col["warning"]

    def test_overview_formula_column_is_classified_not_executed(self, table_tmp_path):
        pytest.importorskip("openpyxl")
        from openpyxl import Workbook

        p = os.path.join(table_tmp_path, "formula_overview.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(["Score"])
        ws.append(["=SUM(1,2)"])
        wb.save(p)

        result = _call(_rel(p), "overview")
        assert result["status"] == "success"
        col = result["columns"][0]
        assert col["classification"] == "formula-containing"
        assert col.get("formula_count") == 1
        assert "not executed" in col["warning"]

    def test_overview_empty_column(self, table_tmp_path):
        p = os.path.join(table_tmp_path, "empty_col.csv")
        _write_csv(p, [["A", "B"], ["1", ""], ["2", ""]])
        result = _call(_rel(p), "overview")
        b_col = next(c for c in result["columns"] if c["column_name"] == "B")
        assert b_col["classification"] == "empty"

    def test_overview_multi_sheet_requires_selection(self, table_tmp_path):
        pytest.importorskip("openpyxl")
        from openpyxl import Workbook

        p = os.path.join(table_tmp_path, "multi_sheet_overview.xlsx")
        wb = Workbook()
        wb.active.title = "A"
        wb.create_sheet("B")
        wb.save(p)

        result = _call(_rel(p), "overview")
        assert result["status"] == "ambiguous"
        assert result["status_reason"] == "multiple_sheets_require_selection"

    def test_overview_preserves_existing_operations(self, table_tmp_path):
        p = os.path.join(table_tmp_path, "scores.csv")
        _write_csv(p, [["Score"], ["10"], ["20"], ["30"]])
        sum_result = _call(_rel(p), "sum", "Score")
        assert sum_result["status"] == "success"
        assert sum_result["computed_value"] == "60"
        max_result = _call(_rel(p), "max", "Score")
        assert max_result["status"] == "success"
        assert max_result["computed_value"] == "30"

    def test_overview_xlsx_single_sheet(self, table_tmp_path):
        pytest.importorskip("openpyxl")
        from openpyxl import Workbook

        p = os.path.join(table_tmp_path, "overview.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "Grades"
        ws.append(["Name", "Score"])
        ws.append(["Alice", 84])
        ws.append(["Bob", 91])
        wb.save(p)

        result = _call(_rel(p), "overview")
        assert result["status"] == "success"
        assert result["sheet_name"] == "Grades"
        assert result["data_row_count"] == 2


    def test_date_cell_is_non_numeric(self, table_tmp_path):
        pytest.importorskip("openpyxl")
        from openpyxl import Workbook
        from datetime import datetime

        p = os.path.join(table_tmp_path, "dates.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.append(["When"])
        ws.append([datetime(2026, 1, 1)])
        wb.save(p)

        result = _call(_rel(p), "max", "When")
        assert result["status"] == "unsupported"
        assert result["status_reason"] == "non_numeric_value_present"

    def test_boolean_cell_is_non_numeric(self, table_tmp_path):
        pytest.importorskip("openpyxl")
        from openpyxl import Workbook

        p = os.path.join(table_tmp_path, "bool.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.append(["Flag"])
        ws.append([True])
        wb.save(p)

        result = _call(_rel(p), "max", "Flag")
        assert result["status"] == "unsupported"
        assert result["status_reason"] == "non_numeric_value_present"


# ---------------------------------------------------------------------------
# F5B-1 Filter operation tests
# ---------------------------------------------------------------------------

def _filter_call(tmp_path, rows, operation="filter", target_column=None, filter_op=None,
                 filter_value=None, filter_value_to=None, filename="data.csv"):
    p = os.path.join(tmp_path, filename)
    _write_csv(p, rows)
    return analyze_table._analyze_table_impl(
        _rel(p), operation,
        target_column=target_column,
        filter_op=filter_op,
        filter_value=filter_value,
        filter_value_to=filter_value_to,
    )


class TestFilterTextOperators:
    def test_eq_text(self, table_tmp_path):
        r = _filter_call(table_tmp_path,
                         [["Name", "Team"], ["Alice", "Blue"], ["Bob", "Red"], ["Cara", "Blue"]],
                         target_column="Team", filter_op="eq", filter_value="Blue")
        assert r["status"] == "success"
        assert r["operation"] == "filter"
        assert r["matched_row_count"] == 2
        assert r["returned_row_count"] == 2
        assert r["result_complete"] is True
        assert r["truncated"] is False
        assert len(r["rows"]) == 2
        assert r["predicate"]["operator"] == "eq"
        assert r["predicate"]["comparison_type"] == "text"

    def test_eq_case_insensitive(self, table_tmp_path):
        r = _filter_call(table_tmp_path,
                         [["Team"], ["blue"], ["BLUE"], ["Blue"], ["red"]],
                         target_column="Team", filter_op="eq", filter_value="blue")
        assert r["status"] == "success"
        assert r["matched_row_count"] == 3

    def test_neq_text(self, table_tmp_path):
        r = _filter_call(table_tmp_path,
                         [["Team"], ["Blue"], ["Red"], ["Blue"]],
                         target_column="Team", filter_op="neq", filter_value="Blue")
        assert r["status"] == "success"
        assert r["matched_row_count"] == 1

    def test_contains(self, table_tmp_path):
        r = _filter_call(table_tmp_path,
                         [["Notes"], ["high risk item"], ["low item"], ["critical risk"]],
                         target_column="Notes", filter_op="contains", filter_value="risk")
        assert r["status"] == "success"
        assert r["matched_row_count"] == 2

    def test_not_contains(self, table_tmp_path):
        r = _filter_call(table_tmp_path,
                         [["Notes"], ["high risk item"], ["low item"], ["critical risk"]],
                         target_column="Notes", filter_op="not_contains", filter_value="risk")
        assert r["status"] == "success"
        assert r["matched_row_count"] == 1

    def test_starts_with(self, table_tmp_path):
        r = _filter_call(table_tmp_path,
                         [["Name"], ["Alice"], ["Aaron"], ["Bob"]],
                         target_column="Name", filter_op="starts_with", filter_value="A")
        assert r["status"] == "success"
        assert r["matched_row_count"] == 2

    def test_ends_with(self, table_tmp_path):
        r = _filter_call(table_tmp_path,
                         [["Code"], ["A99"], ["B99"], ["C00"]],
                         target_column="Code", filter_op="ends_with", filter_value="99")
        assert r["status"] == "success"
        assert r["matched_row_count"] == 2

    def test_is_blank(self, table_tmp_path):
        r = _filter_call(table_tmp_path,
                         [["Notes"], ["present"], [""], [None], ["   "]],
                         target_column="Notes", filter_op="is_blank")
        assert r["status"] == "success"
        assert r["matched_row_count"] == 3  # "", None, "   "

    def test_is_not_blank(self, table_tmp_path):
        r = _filter_call(table_tmp_path,
                         [["Notes"], ["present"], [""], [None]],
                         target_column="Notes", filter_op="is_not_blank")
        assert r["status"] == "success"
        assert r["matched_row_count"] == 1

    def test_is_blank_does_not_treat_zero_as_blank(self, table_tmp_path):
        r = _filter_call(table_tmp_path,
                         [["Score"], ["0"], ["1"]],
                         target_column="Score", filter_op="is_blank")
        assert r["status"] == "success"
        assert r["matched_row_count"] == 0


class TestFilterNumericOperators:
    def test_numeric_eq(self, table_tmp_path):
        r = _filter_call(table_tmp_path,
                         [["Score"], ["85"], ["91"], ["85"]],
                         target_column="Score", filter_op="eq", filter_value="85")
        assert r["status"] == "success"
        assert r["matched_row_count"] == 2
        assert r["predicate"]["comparison_type"] == "numeric"

    def test_numeric_neq(self, table_tmp_path):
        r = _filter_call(table_tmp_path,
                         [["Score"], ["85"], ["91"], ["85"]],
                         target_column="Score", filter_op="neq", filter_value="85")
        assert r["status"] == "success"
        assert r["matched_row_count"] == 1

    def test_gt(self, table_tmp_path):
        r = _filter_call(table_tmp_path,
                         [["Score"], ["80"], ["85"], ["91"]],
                         target_column="Score", filter_op="gt", filter_value="85")
        assert r["status"] == "success"
        assert r["matched_row_count"] == 1

    def test_gte(self, table_tmp_path):
        r = _filter_call(table_tmp_path,
                         [["Score"], ["80"], ["85"], ["91"]],
                         target_column="Score", filter_op="gte", filter_value="85")
        assert r["status"] == "success"
        assert r["matched_row_count"] == 2

    def test_lt(self, table_tmp_path):
        r = _filter_call(table_tmp_path,
                         [["Score"], ["80"], ["85"], ["91"]],
                         target_column="Score", filter_op="lt", filter_value="85")
        assert r["status"] == "success"
        assert r["matched_row_count"] == 1

    def test_lte(self, table_tmp_path):
        r = _filter_call(table_tmp_path,
                         [["Score"], ["80"], ["85"], ["91"]],
                         target_column="Score", filter_op="lte", filter_value="85")
        assert r["status"] == "success"
        assert r["matched_row_count"] == 2

    def test_between_inclusive(self, table_tmp_path):
        r = _filter_call(table_tmp_path,
                         [["Score"], ["79"], ["80"], ["85"], ["90"], ["91"]],
                         target_column="Score", filter_op="between",
                         filter_value="80", filter_value_to="90")
        assert r["status"] == "success"
        assert r["matched_row_count"] == 3  # 80, 85, 90

    def test_between_invalid_range(self, table_tmp_path):
        r = _filter_call(table_tmp_path,
                         [["Score"], ["80"], ["85"]],
                         target_column="Score", filter_op="between",
                         filter_value="90", filter_value_to="80")
        assert r["status"] == "unsupported"
        assert r["status_reason"] == "invalid_filter_range"

    def test_filter_value_not_numeric(self, table_tmp_path):
        r = _filter_call(table_tmp_path,
                         [["Score"], ["80"], ["85"]],
                         target_column="Score", filter_op="gt", filter_value="abc")
        assert r["status"] == "unsupported"
        assert r["status_reason"] == "filter_value_not_numeric"

    def test_filter_upper_value_not_numeric(self, table_tmp_path):
        r = _filter_call(table_tmp_path,
                         [["Score"], ["80"], ["85"]],
                         target_column="Score", filter_op="between",
                         filter_value="80", filter_value_to="abc")
        assert r["status"] == "unsupported"
        assert r["status_reason"] == "filter_upper_value_not_numeric"


class TestFilterSemantics:
    def test_zero_matches_is_success(self, table_tmp_path):
        r = _filter_call(table_tmp_path,
                         [["Team"], ["Blue"], ["Red"]],
                         target_column="Team", filter_op="eq", filter_value="Green")
        assert r["status"] == "success"
        assert r["matched_row_count"] == 0
        assert r["returned_row_count"] == 0
        assert r["result_complete"] is True
        assert r["rows"] == []
        assert "No rows match" in r["answer_text"]

    def test_all_rows_match(self, table_tmp_path):
        r = _filter_call(table_tmp_path,
                         [["Score"], ["10"], ["20"], ["30"]],
                         target_column="Score", filter_op="gt", filter_value="5")
        assert r["status"] == "success"
        assert r["matched_row_count"] == 3
        assert r["result_complete"] is True

    def test_exact_matched_row_count_with_truncation(self, table_tmp_path):
        rows = [["Score"]] + [[str(i)] for i in range(1100)]
        r = _filter_call(table_tmp_path, rows,
                         target_column="Score", filter_op="gte", filter_value="0")
        assert r["status"] == "success"
        assert r["matched_row_count"] == 1100
        assert r["returned_row_count"] == 1000
        assert r["result_complete"] is False
        assert r["truncated"] is True
        assert len(r["rows"]) == 1000
        assert "1,100" in r["answer_text"] or "1100" in r["answer_text"]
        assert len(r["limitations"]) > 0

    def test_original_source_order_preserved(self, table_tmp_path):
        r = _filter_call(table_tmp_path,
                         [["Name", "Score"],
                          ["Alice", "90"], ["Bob", "70"], ["Cara", "85"]],
                         target_column="Score", filter_op="gte", filter_value="85")
        assert r["status"] == "success"
        names = [c["value"] for row in r["rows"] for c in row["cells"] if c["column_name"] == "Name"]
        assert names == ["Alice", "Cara"]

    def test_row_serialization_a1_refs(self, table_tmp_path):
        r = _filter_call(table_tmp_path,
                         [["Name", "Score"], ["Alice", "90"], ["Bob", "70"]],
                         target_column="Score", filter_op="eq", filter_value="90")
        assert r["status"] == "success"
        assert len(r["rows"]) == 1
        row = r["rows"][0]
        assert row["row_number"] == 1
        assert row["row_ref"] == "row:1"
        assert len(row["cells"]) == 2
        name_cell = next(c for c in row["cells"] if c["column_name"] == "Name")
        assert name_cell["cell_ref"] == "A2"
        assert name_cell["column_index"] == 1
        score_cell = next(c for c in row["cells"] if c["column_name"] == "Score")
        assert score_cell["cell_ref"] == "B2"

    def test_answer_text_small_complete_with_names(self, table_tmp_path):
        r = _filter_call(table_tmp_path,
                         [["Name", "Team"], ["Alice", "Blue"], ["Bob", "Red"], ["Cara", "Blue"]],
                         target_column="Team", filter_op="eq", filter_value="Blue")
        assert r["status"] == "success"
        assert "Alice" in r["answer_text"] or "Cara" in r["answer_text"]

    def test_answer_text_zero_match(self, table_tmp_path):
        r = _filter_call(table_tmp_path,
                         [["Team"], ["Blue"]],
                         target_column="Team", filter_op="eq", filter_value="Green")
        assert "No rows match" in r["answer_text"]

    def test_text_eq_is_exact_casefold_not_numeric_parse(self, table_tmp_path):
        r = _filter_call(table_tmp_path,
                         [["Code"], ["A001"], ["A002"], ["A003"]],
                         target_column="Code", filter_op="eq", filter_value="A001")
        assert r["status"] == "success"
        assert r["matched_row_count"] == 1
        # Original value in serialized row must be "A001" (unchanged)
        assert r["rows"][0]["cells"][0]["value"] == "A001"

    def test_text_eq_not_substring(self, table_tmp_path):
        r = _filter_call(table_tmp_path,
                         [["Code"], ["A001"], ["A002"], ["A003"]],
                         target_column="Code", filter_op="eq", filter_value="A")
        assert r["status"] == "success"
        assert r["matched_row_count"] == 0

    def test_filter_value_too_long(self, table_tmp_path):
        r = _filter_call(table_tmp_path,
                         [["Name"], ["Alice"]],
                         target_column="Name", filter_op="eq",
                         filter_value="x" * 501)
        assert r["status"] == "unsupported"
        assert r["status_reason"] == "filter_value_too_long"

    def test_file_unmodified(self, table_tmp_path):
        p = os.path.join(table_tmp_path, "data.csv")
        _write_csv(p, [["Name"], ["Alice"]])
        import hashlib
        with open(p, "rb") as f:
            before = hashlib.md5(f.read()).hexdigest()
        _filter_call(table_tmp_path, [["Name"], ["Alice"]],
                     target_column="Name", filter_op="eq", filter_value="Alice")
        with open(p, "rb") as f:
            after = hashlib.md5(f.read()).hexdigest()
        assert before == after


class TestFilterErrors:
    def test_missing_column(self, table_tmp_path):
        r = _filter_call(table_tmp_path,
                         [["Name"], ["Alice"]],
                         target_column="Score", filter_op="eq", filter_value="90")
        assert r["status"] == "not_found"
        assert r["status_reason"] == "column_not_found"

    def test_duplicate_header_ambiguity(self, table_tmp_path):
        r = _filter_call(table_tmp_path,
                         [["Score", "Score"], ["90", "80"]],
                         target_column="Score", filter_op="eq", filter_value="90")
        assert r["status"] == "ambiguous"
        assert r["status_reason"] == "duplicate_column_header"

    def test_formula_target_rejected(self, table_tmp_path):
        pytest.importorskip("openpyxl")
        from openpyxl import Workbook
        p = os.path.join(table_tmp_path, "formula.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.append(["Score"])
        ws.append(["=A1+1"])
        wb.save(p)
        r = analyze_table._analyze_table_impl(
            _rel(p), "filter",
            target_column="Score",
            filter_op="gt",
            filter_value="0",
        )
        assert r["status"] == "unsupported"
        assert r["status_reason"] == "formula_cell_present"

    def test_mixed_target_text_filter_works(self, table_tmp_path):
        r = _filter_call(table_tmp_path,
                         [["Status"], ["Active"], ["42"], ["Inactive"]],
                         target_column="Status", filter_op="eq", filter_value="Active")
        assert r["status"] == "success"
        assert r["matched_row_count"] == 1
        assert r["predicate"]["comparison_type"] == "text"
        assert r["warnings"]  # warning about mixed column

    def test_mixed_target_numeric_operator_rejected(self, table_tmp_path):
        r = _filter_call(table_tmp_path,
                         [["Status"], ["Active"], ["42"]],
                         target_column="Status", filter_op="gt", filter_value="10")
        assert r["status"] == "unsupported"
        assert r["status_reason"] == "column_type_mismatch"

    def test_numeric_operator_on_text_column_rejected(self, table_tmp_path):
        r = _filter_call(table_tmp_path,
                         [["Team"], ["Blue"], ["Red"]],
                         target_column="Team", filter_op="gt", filter_value="5")
        assert r["status"] == "unsupported"
        assert r["status_reason"] == "column_type_mismatch"

    def test_contains_on_numeric_column_rejected(self, table_tmp_path):
        r = _filter_call(table_tmp_path,
                         [["Score"], ["80"], ["85"]],
                         target_column="Score", filter_op="contains", filter_value="8")
        assert r["status"] == "unsupported"
        assert r["status_reason"] == "operator_not_supported_for_numeric_column"


# ---------------------------------------------------------------------------
# FIX A — run() wrapper: controlled domain result envelope semantics
# ---------------------------------------------------------------------------

def _run_call(tmp_path, rows, operation="filter", target_column=None, filter_op=None,
              filter_value=None, filter_value_to=None, filename="data.csv"):
    """Call the public run() entry point (what governance sees)."""
    p = os.path.join(tmp_path, filename)
    _write_csv(p, rows)
    return analyze_table.run(
        _rel(p), operation,
        target_column=target_column,
        filter_op=filter_op,
        filter_value=filter_value,
        filter_value_to=filter_value_to,
    )


class TestControlledDomainResultEnvelope:
    """
    FIX B: Deterministic terminal domain outcomes must complete the workflow once.
    run() must return outer status=success with inner structured result for
    unsupported/not_found/ambiguous outcomes so governance does not retry them.
    """

    def test_column_type_mismatch_is_outer_success(self, table_tmp_path):
        r = _run_call(table_tmp_path,
                      [["Notes"], ["hello"], ["world"]],
                      target_column="Notes", filter_op="gt", filter_value="5")
        assert r["status"] == "success", f"Expected outer success, got {r}"
        inner = r["result"]
        assert inner["status"] == "unsupported"
        assert inner["status_reason"] == "column_type_mismatch"
        assert isinstance(inner.get("answer_text"), str)
        assert len(inner["answer_text"]) > 0
        assert "Notes" in inner["answer_text"]

    def test_column_type_mismatch_has_no_retry_trigger(self, table_tmp_path):
        r = _run_call(table_tmp_path,
                      [["Notes"], ["hello"], ["world"]],
                      target_column="Notes", filter_op="gt", filter_value="5")
        assert r["status"] == "success"
        assert r["result"]["status"] == "unsupported"

    def test_operator_not_supported_for_numeric_is_outer_success(self, table_tmp_path):
        r = _run_call(table_tmp_path,
                      [["Score"], ["80"], ["90"]],
                      target_column="Score", filter_op="contains", filter_value="8")
        assert r["status"] == "success"
        inner = r["result"]
        assert inner["status"] == "unsupported"
        assert inner["status_reason"] == "operator_not_supported_for_numeric_column"
        assert isinstance(inner.get("answer_text"), str)

    def test_formula_cell_present_is_outer_success(self, table_tmp_path):
        pytest.importorskip("openpyxl")
        from openpyxl import Workbook
        p = os.path.join(table_tmp_path, "formula_run.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.append(["Score"])
        ws.append(["=A1+1"])
        wb.save(p)
        r = analyze_table.run(
            _rel(p), "filter",
            target_column="Score",
            filter_op="gt",
            filter_value="0",
        )
        assert r["status"] == "success"
        assert r["result"]["status"] == "unsupported"
        assert r["result"]["status_reason"] == "formula_cell_present"
        assert isinstance(r["result"].get("answer_text"), str)

    def test_column_not_found_is_outer_success(self, table_tmp_path):
        r = _run_call(table_tmp_path,
                      [["Name"], ["Alice"]],
                      target_column="NonExistentColumn", filter_op="eq", filter_value="x")
        assert r["status"] == "success"
        inner = r["result"]
        assert inner["status"] == "not_found"
        assert inner["status_reason"] == "column_not_found"
        assert isinstance(inner.get("answer_text"), str)

    def test_duplicate_column_header_is_outer_success(self, table_tmp_path):
        r = _run_call(table_tmp_path,
                      [["Score", "Score"], ["90", "80"]],
                      target_column="Score", filter_op="eq", filter_value="90")
        assert r["status"] == "success"
        inner = r["result"]
        assert inner["status"] == "ambiguous"
        assert inner["status_reason"] == "duplicate_column_header"
        assert isinstance(inner.get("answer_text"), str)

    def test_between_on_text_column_is_outer_success(self, table_tmp_path):
        r = _run_call(table_tmp_path,
                      [["Notes"], ["alpha"], ["beta"]],
                      target_column="Notes", filter_op="between",
                      filter_value="a", filter_value_to="z")
        assert r["status"] == "success"
        inner = r["result"]
        assert inner["status"] == "unsupported"
        assert inner["status_reason"] == "column_type_mismatch"
        assert isinstance(inner.get("answer_text"), str)

    def test_filter_value_not_numeric_is_outer_success(self, table_tmp_path):
        r = _run_call(table_tmp_path,
                      [["Score"], ["80"], ["90"]],
                      target_column="Score", filter_op="gt", filter_value="not-a-number")
        assert r["status"] == "success"
        inner = r["result"]
        assert inner["status"] == "unsupported"
        assert isinstance(inner.get("answer_text"), str)

    def test_file_not_found_is_outer_success(self):
        r = analyze_table.run(
            "tmp/nonexistent_file_xyz.csv", "filter",
            target_column="Score", filter_op="eq", filter_value="x",
        )
        assert r["status"] == "success"
        assert r["result"]["status"] == "not_found"
        assert r["result"]["status_reason"] == "file_not_found"
        assert isinstance(r["result"].get("answer_text"), str)

    def test_genuine_failure_remains_outer_failure(self, tmp_path):
        r = analyze_table.run(
            str(tmp_path / "impossible" / "blocked.csv"), "filter",
            target_column="Score", filter_op="eq", filter_value="x",
        )
        assert r["status"] in ("failure", "success")
        if r["status"] == "success":
            assert r["result"]["status"] in ("not_found", "blocked", "unsupported")

    def test_successful_filter_outer_success_inner_success(self, table_tmp_path):
        r = _run_call(table_tmp_path,
                      [["Score"], ["80"], ["90"], ["70"]],
                      target_column="Score", filter_op="gte", filter_value="85")
        assert r["status"] == "success"
        assert r["result"]["status"] == "success"
        assert r["result"]["operation"] == "filter"
        assert r["result"]["matched_row_count"] == 1

    def test_f5a_overview_run_succeeds(self, table_tmp_path):
        p = os.path.join(table_tmp_path, "overview_run.csv")
        _write_csv(p, [["Name", "Score"], ["Alice", "90"], ["Bob", "80"]])
        r = analyze_table.run(_rel(p), "overview")
        assert r["status"] == "success"
        assert r["result"]["status"] == "success"
        assert r["result"]["operation"] == "overview"

    def test_f5a_max_run_succeeds(self, table_tmp_path):
        p = os.path.join(table_tmp_path, "max_run.csv")
        _write_csv(p, [["Score"], ["80"], ["90"], ["70"]])
        r = analyze_table.run(_rel(p), "max", target_column="Score")
        assert r["status"] == "success"
        assert r["result"]["status"] == "success"
        assert int(r["result"]["computed_value"]) == 90

    def test_no_observation_key_on_controlled_success(self, table_tmp_path):
        r = _run_call(table_tmp_path,
                      [["Notes"], ["hello"]],
                      target_column="Notes", filter_op="gt", filter_value="5")
        assert r["status"] == "success"
        assert "observation" not in r


class TestFilterXlsx:
    def test_filter_xlsx_eq(self, table_tmp_path):
        pytest.importorskip("openpyxl")
        from openpyxl import Workbook
        p = os.path.join(table_tmp_path, "data.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(["Name", "Team"])
        ws.append(["Alice", "Blue"])
        ws.append(["Bob", "Red"])
        wb.save(p)
        r = analyze_table._analyze_table_impl(
            _rel(p), "filter",
            target_column="Team",
            filter_op="eq",
            filter_value="Blue",
        )
        assert r["status"] == "success"
        assert r["matched_row_count"] == 1
        assert r["sheet_name"] == "Sheet1"

    def test_filter_xlsx_explicit_sheet(self, table_tmp_path):
        pytest.importorskip("openpyxl")
        from openpyxl import Workbook
        p = os.path.join(table_tmp_path, "multi.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "Grades"
        ws.append(["Score"])
        ws.append([90])
        ws.append([70])
        wb.create_sheet("Other")
        wb.save(p)
        r = analyze_table._analyze_table_impl(
            _rel(p), "filter",
            target_column="Score",
            filter_op="gte",
            filter_value="85",
            sheet_name="Grades",
        )
        assert r["status"] == "success"
        assert r["matched_row_count"] == 1
        assert r["sheet_name"] == "Grades"

    def test_filter_xlsx_multi_sheet_ambiguity(self, table_tmp_path):
        pytest.importorskip("openpyxl")
        from openpyxl import Workbook
        p = os.path.join(table_tmp_path, "multi2.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(["Score"])
        ws.append([90])
        wb.create_sheet("Sheet2")
        wb.save(p)
        r = analyze_table._analyze_table_impl(
            _rel(p), "filter",
            target_column="Score",
            filter_op="eq",
            filter_value="90",
        )
        assert r["status"] == "ambiguous"
        assert r["status_reason"] == "multiple_sheets_require_selection"


# ---------------------------------------------------------------------------
# F5R: trust_metadata on run() results
# ---------------------------------------------------------------------------

_REQUIRED_TRUST_FIELDS = [
    "trust_class", "verification_status", "plan_version", "plan_source_path",
    "requested_operations", "executed_operations", "omitted_operations",
    "operation_coverage_complete", "result_complete", "evidence_refs",
    "source_context_refs", "context_scope", "context_complete",
    "advisory_disclaimer", "unsupported_reason", "ambiguity_reason",
    "clarification_needed", "limitations", "warnings",
    "learning_eligible", "operator_acceptance_status",
]


class TestTrustMetadataOnRunSuccess:
    """run() must attach trust_metadata with trust_class=verified to every success result."""

    def test_count_rows_has_trust_metadata(self, table_tmp_path):
        p = os.path.join(table_tmp_path, "scores.csv")
        _write_csv(p, [["Score"], ["10"], ["20"]])
        r = analyze_table.run(_rel(p), "count_rows")
        assert r["status"] == "success"
        meta = r["result"].get("trust_metadata")
        assert meta is not None, "trust_metadata missing from count_rows result"
        for field in _REQUIRED_TRUST_FIELDS:
            assert field in meta, f"trust_metadata missing field: {field}"

    def test_count_rows_trust_class_verified(self, table_tmp_path):
        p = os.path.join(table_tmp_path, "scores.csv")
        _write_csv(p, [["Score"], ["10"], ["20"]])
        r = analyze_table.run(_rel(p), "count_rows")
        meta = r["result"]["trust_metadata"]
        assert meta["trust_class"] == "verified"
        assert meta["verification_status"] == "verified"
        assert meta["operation_coverage_complete"] is True

    def test_max_trust_verified_with_evidence(self, table_tmp_path):
        p = os.path.join(table_tmp_path, "scores.csv")
        _write_csv(p, [["Name", "Score"], ["Alice", "90"], ["Bob", "80"]])
        r = analyze_table.run(_rel(p), "max", target_column="Score")
        meta = r["result"]["trust_metadata"]
        assert meta["trust_class"] == "verified"
        assert len(meta["evidence_refs"]) > 0

    def test_filter_success_trust_verified(self, table_tmp_path):
        p = os.path.join(table_tmp_path, "scores.csv")
        _write_csv(p, [["Score"], ["80"], ["90"]])
        r = analyze_table.run(_rel(p), "filter", target_column="Score",
                              filter_op="gte", filter_value="85")
        assert r["status"] == "success"
        meta = r["result"]["trust_metadata"]
        assert meta["trust_class"] == "verified"

    def test_overview_trust_verified(self, table_tmp_path):
        p = os.path.join(table_tmp_path, "overview.csv")
        _write_csv(p, [["Name", "Score"], ["Alice", "90"], ["Bob", "80"]])
        r = analyze_table.run(_rel(p), "overview")
        meta = r["result"]["trust_metadata"]
        assert meta["trust_class"] == "verified"
        assert meta["learning_eligible"] is False

    def test_trust_metadata_learning_always_false(self, table_tmp_path):
        p = os.path.join(table_tmp_path, "scores.csv")
        _write_csv(p, [["Score"], ["10"]])
        r = analyze_table.run(_rel(p), "sum", target_column="Score")
        assert r["result"]["trust_metadata"]["learning_eligible"] is False

    def test_trust_metadata_operator_acceptance_status(self, table_tmp_path):
        p = os.path.join(table_tmp_path, "scores.csv")
        _write_csv(p, [["Score"], ["10"]])
        r = analyze_table.run(_rel(p), "count_rows")
        assert r["result"]["trust_metadata"]["operator_acceptance_status"] == "unreviewed"


class TestTrustMetadataOnControlledOutcomes:
    """run() must attach trust_metadata with trust_class=unsupported or ambiguous
    to controlled domain outcomes (still wrapped in outer success)."""

    def test_column_type_mismatch_trust_unsupported(self, table_tmp_path):
        p = os.path.join(table_tmp_path, "notes.csv")
        _write_csv(p, [["Notes"], ["hello"], ["world"]])
        r = analyze_table.run(_rel(p), "filter", target_column="Notes",
                              filter_op="gt", filter_value="5")
        assert r["status"] == "success"
        meta = r["result"].get("trust_metadata")
        assert meta is not None
        assert meta["trust_class"] == "unsupported"
        assert meta["learning_eligible"] is False

    def test_formula_cell_trust_unsupported(self, table_tmp_path):
        pytest.importorskip("openpyxl")
        from openpyxl import Workbook
        p = os.path.join(table_tmp_path, "formula.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.append(["Score"])
        ws.append(["=A1+1"])
        wb.save(p)
        r = analyze_table.run(_rel(p), "filter", target_column="Score",
                              filter_op="gt", filter_value="0")
        assert r["status"] == "success"
        meta = r["result"].get("trust_metadata")
        assert meta is not None
        assert meta["trust_class"] == "unsupported"

    def test_column_not_found_trust_ambiguous(self, table_tmp_path):
        p = os.path.join(table_tmp_path, "names.csv")
        _write_csv(p, [["Name"], ["Alice"]])
        r = analyze_table.run(_rel(p), "filter", target_column="NonExistent",
                              filter_op="eq", filter_value="Alice")
        assert r["status"] == "success"
        meta = r["result"].get("trust_metadata")
        assert meta is not None
        assert meta["trust_class"] == "ambiguous"
        assert meta["clarification_needed"] is True

    def test_duplicate_column_trust_ambiguous(self, table_tmp_path):
        p = os.path.join(table_tmp_path, "dup.csv")
        _write_csv(p, [["Score", "Score"], ["90", "80"]])
        r = analyze_table.run(_rel(p), "filter", target_column="Score",
                              filter_op="eq", filter_value="90")
        assert r["status"] == "success"
        meta = r["result"].get("trust_metadata")
        assert meta is not None
        assert meta["trust_class"] == "ambiguous"

    def test_controlled_outcome_has_all_required_fields(self, table_tmp_path):
        p = os.path.join(table_tmp_path, "notes.csv")
        _write_csv(p, [["Notes"], ["hello"]])
        r = analyze_table.run(_rel(p), "filter", target_column="Notes",
                              filter_op="gt", filter_value="5")
        meta = r["result"].get("trust_metadata", {})
        for field in _REQUIRED_TRUST_FIELDS:
            assert field in meta, f"trust_metadata missing field: {field}"


# ---------------------------------------------------------------------------
# F5R: run_plan() multi-filter AND semantics
# ---------------------------------------------------------------------------

def _make_multi_filter_plan(tmp_path, filters, source_filename="data.csv"):
    """Build a valid TableAnalysisPlanV1 pointing at a temp CSV."""
    from system.orchestrator.structured_data.table_analysis_plan import (
        build_multi_filter_sort_plan,
        PLAN_VERSION,
        MAX_OPERATIONS,
        MAX_PREDICATES,
        MAX_ROWS_SCANNED,
        MAX_ROWS_RETURNED,
    )
    path = os.path.join(tmp_path, source_filename)
    rel_path = _rel(path)
    return build_multi_filter_sort_plan(rel_path, filters), path, rel_path


class TestRunPlanMultiFilter:
    def test_two_filter_and_semantics(self, table_tmp_path):
        p = os.path.join(table_tmp_path, "data.csv")
        _write_csv(p, [
            ["Name", "Score", "Team"],
            ["Alice", "90", "Blue"],
            ["Bob", "90", "Red"],
            ["Cara", "80", "Blue"],
            ["Dave", "70", "Red"],
        ])
        from system.orchestrator.structured_data.table_analysis_plan import build_multi_filter_sort_plan
        filters = [
            {"column": "Score", "filter_op": "gte", "filter_value": "85"},
            {"column": "Team", "filter_op": "eq", "filter_value": "Blue"},
        ]
        plan = build_multi_filter_sort_plan(_rel(p), filters)
        r = analyze_table.run_plan(plan)
        assert r["status"] == "success", f"run_plan failed: {r}"
        result = r["result"]
        assert result["matched_row_count"] == 1
        names = [c["value"] for row in result["rows"]
                 for c in row["cells"] if c["column_name"] == "Name"]
        assert names == ["Alice"]

    def test_two_filters_zero_matches(self, table_tmp_path):
        p = os.path.join(table_tmp_path, "data.csv")
        _write_csv(p, [
            ["Score", "Team"],
            ["90", "Blue"],
            ["80", "Red"],
        ])
        from system.orchestrator.structured_data.table_analysis_plan import build_multi_filter_sort_plan
        filters = [
            {"column": "Score", "filter_op": "gte", "filter_value": "90"},
            {"column": "Team", "filter_op": "eq", "filter_value": "Red"},
        ]
        plan = build_multi_filter_sort_plan(_rel(p), filters)
        r = analyze_table.run_plan(plan)
        assert r["status"] == "success"
        assert r["result"]["matched_row_count"] == 0
        assert "No rows match" in r["result"]["answer_text"]

    def test_multi_filter_trust_verified(self, table_tmp_path):
        p = os.path.join(table_tmp_path, "data.csv")
        _write_csv(p, [
            ["Score", "Team"],
            ["90", "Blue"],
            ["80", "Blue"],
            ["70", "Red"],
        ])
        from system.orchestrator.structured_data.table_analysis_plan import build_multi_filter_sort_plan
        filters = [
            {"column": "Team", "filter_op": "eq", "filter_value": "Blue"},
            {"column": "Score", "filter_op": "gte", "filter_value": "85"},
        ]
        plan = build_multi_filter_sort_plan(_rel(p), filters)
        r = analyze_table.run_plan(plan)
        assert r["status"] == "success"
        meta = r["result"]["trust_metadata"]
        assert meta["trust_class"] == "verified"
        assert meta["operation_coverage_complete"] is True

    def test_multi_filter_trust_metadata_has_all_fields(self, table_tmp_path):
        p = os.path.join(table_tmp_path, "data.csv")
        _write_csv(p, [["Score", "Team"], ["90", "Blue"]])
        from system.orchestrator.structured_data.table_analysis_plan import build_multi_filter_sort_plan
        filters = [
            {"column": "Score", "filter_op": "gt", "filter_value": "80"},
        ]
        plan = build_multi_filter_sort_plan(_rel(p), filters)
        r = analyze_table.run_plan(plan)
        assert r["status"] == "success"
        meta = r["result"].get("trust_metadata", {})
        for field in _REQUIRED_TRUST_FIELDS:
            assert field in meta, f"trust_metadata missing field: {field}"

    def test_run_plan_invalid_plan_returns_failure(self):
        r = analyze_table.run_plan({"version": "wrong"})
        assert r["status"] == "failure"
        assert "plan_validation_failed" in r["reason"]

    def test_run_plan_single_op_delegates_to_legacy(self, table_tmp_path):
        p = os.path.join(table_tmp_path, "data.csv")
        _write_csv(p, [["Score"], ["10"], ["20"], ["30"]])
        from system.orchestrator.structured_data.table_analysis_plan import build_single_op_plan
        plan = build_single_op_plan(_rel(p), "count_rows")
        r = analyze_table.run_plan(plan)
        assert r["status"] == "success"
        assert r["result"]["operation"] == "count_rows"
        assert r["result"]["trust_metadata"]["trust_class"] == "verified"


class TestRunPlanSort:
    def test_filter_then_sort_asc(self, table_tmp_path):
        p = os.path.join(table_tmp_path, "data.csv")
        _write_csv(p, [
            ["Name", "Score"],
            ["Cara", "85"],
            ["Alice", "90"],
            ["Bob", "85"],
            ["Dave", "70"],
        ])
        from system.orchestrator.structured_data.table_analysis_plan import build_multi_filter_sort_plan
        filters = [{"column": "Score", "filter_op": "gte", "filter_value": "85"}]
        plan = build_multi_filter_sort_plan(_rel(p), filters, sort_column="Name", sort_direction="asc")
        r = analyze_table.run_plan(plan)
        assert r["status"] == "success"
        result = r["result"]
        assert result["matched_row_count"] == 3
        names = [c["value"] for row in result["rows"]
                 for c in row["cells"] if c["column_name"] == "Name"]
        assert names == ["Alice", "Bob", "Cara"]

    def test_filter_then_sort_desc(self, table_tmp_path):
        p = os.path.join(table_tmp_path, "data.csv")
        _write_csv(p, [
            ["Name", "Score"],
            ["Alice", "90"],
            ["Bob", "85"],
            ["Cara", "95"],
        ])
        from system.orchestrator.structured_data.table_analysis_plan import build_multi_filter_sort_plan
        filters = [{"column": "Score", "filter_op": "gt", "filter_value": "80"}]
        plan = build_multi_filter_sort_plan(_rel(p), filters, sort_column="Score", sort_direction="desc")
        r = analyze_table.run_plan(plan)
        assert r["status"] == "success"
        result = r["result"]
        scores = [c["value"] for row in result["rows"]
                  for c in row["cells"] if c["column_name"] == "Score"]
        assert scores == ["95", "90", "85"]

    def test_sort_blanks_always_last(self, table_tmp_path):
        p = os.path.join(table_tmp_path, "data.csv")
        _write_csv(p, [
            ["Name", "Score"],
            ["Alice", "90"],
            ["Bob", ""],
            ["Cara", "85"],
        ])
        from system.orchestrator.structured_data.table_analysis_plan import build_multi_filter_sort_plan
        filters = [{"column": "Name", "filter_op": "is_not_blank"}]
        plan = build_multi_filter_sort_plan(_rel(p), filters, sort_column="Score", sort_direction="asc")
        r = analyze_table.run_plan(plan)
        assert r["status"] == "success"
        scores = [c["value"] for row in r["result"]["rows"]
                  for c in row["cells"] if c["column_name"] == "Score"]
        non_blank_scores = [s for s in scores if s and str(s).strip()]
        assert non_blank_scores == sorted(non_blank_scores, key=lambda x: float(x))
        blank_scores = [s for s in scores if not s or not str(s).strip()]
        assert all(i >= len(non_blank_scores) for i, s in enumerate(scores) if not s or not str(s).strip())

    def test_sort_mixed_type_column_rejected(self, table_tmp_path):
        p = os.path.join(table_tmp_path, "data.csv")
        _write_csv(p, [
            ["Status", "Score"],
            ["Active", "90"],
            ["42", "80"],
        ])
        from system.orchestrator.structured_data.table_analysis_plan import build_multi_filter_sort_plan
        filters = [{"column": "Score", "filter_op": "gt", "filter_value": "70"}]
        plan = build_multi_filter_sort_plan(_rel(p), filters, sort_column="Status", sort_direction="asc")
        r = analyze_table.run_plan(plan)
        # Mixed-type sort is a controlled domain outcome: outer success, inner unsupported
        assert r["status"] == "success"
        assert r["result"]["status"] == "unsupported"
        assert r["result"]["status_reason"] == "sort_mixed_type_column"

    def test_sort_result_operation_id_in_executed(self, table_tmp_path):
        p = os.path.join(table_tmp_path, "data.csv")
        _write_csv(p, [["Name", "Score"], ["Alice", "90"], ["Bob", "80"]])
        from system.orchestrator.structured_data.table_analysis_plan import build_multi_filter_sort_plan
        filters = [{"column": "Score", "filter_op": "gt", "filter_value": "70"}]
        plan = build_multi_filter_sort_plan(_rel(p), filters, sort_column="Name", sort_direction="asc")
        r = analyze_table.run_plan(plan)
        assert r["status"] == "success"
        meta = r["result"]["trust_metadata"]
        assert "op_sort_1" in meta["executed_operations"]


# ---------------------------------------------------------------------------
# F5R: missing-source regression (trust_class=ambiguous, no retry)
# ---------------------------------------------------------------------------

class TestMissingSourceRegression:
    """Regression guard: filter with no source path must be caught at the
    document_local_read capability layer as ambiguous. analyze_table itself
    must return a blocked/not_found result if called directly with no path."""

    def test_analyze_table_run_empty_path_returns_terminal_not_retry(self):
        """Direct call to run() with empty path must return a terminal outcome — no retry.
        Empty path → file not found → controlled domain outcome (outer success, inner not_found)."""
        r = analyze_table.run("", "filter", target_column="Score",
                              filter_op="eq", filter_value="test")
        # Must be a terminal outcome — either outer failure or controlled domain success
        assert r["status"] in ("failure", "success")
        if r["status"] == "success":
            assert r["result"]["status"] in ("not_found", "blocked", "unsupported")

    def test_analyze_table_run_none_path_returns_terminal(self):
        r = analyze_table.run(None, "filter", target_column="Score",
                              filter_op="eq", filter_value="test")
        assert r["status"] in ("failure", "success")
        if r["status"] == "success":
            assert r["result"]["status"] in ("not_found", "blocked", "unsupported", "ambiguous")

    def test_document_local_read_filter_without_path_routes_to_guidance(self):
        """Integration: filter grammar without path → guidance workflow (not analyze_table call)."""
        try:
            from system.orchestrator.capabilities.document_local_read_capability import (
                compile_document_local_read_workflow,
            )
        except ImportError:
            pytest.skip("document_local_read_capability not importable")

        result = compile_document_local_read_workflow(
            "show rows where Score is greater than 85"
        )
        assert result is not None, "Expected guidance workflow, got None"
        assert result.get("steps"), "Expected at least one step"
        step = result["steps"][0]
        meta = step.get("capability_metadata", {})
        assert meta.get("route_reason_code") == "missing_path_filter_guidance"

    def test_missing_path_workflow_has_trust_metadata_ambiguous(self):
        """Guidance workflow must carry trust_metadata with trust_class=ambiguous."""
        try:
            from system.orchestrator.capabilities.document_local_read_capability import (
                compile_document_local_read_workflow,
            )
        except ImportError:
            pytest.skip("document_local_read_capability not importable")

        result = compile_document_local_read_workflow(
            "show rows where Score is greater than 85"
        )
        step = result["steps"][0]
        trust = step["capability_metadata"].get("trust_metadata", {})
        assert trust.get("trust_class") == "ambiguous"
        assert trust.get("ambiguity_reason") == "missing_source_path"
        assert trust.get("clarification_needed") is True
        assert trust.get("learning_eligible") is False

    def test_missing_path_workflow_no_tool_call(self):
        """Guidance workflow step must NOT have a tool_call (no analyze_table invocation)."""
        try:
            from system.orchestrator.capabilities.document_local_read_capability import (
                compile_document_local_read_workflow,
            )
        except ImportError:
            pytest.skip("document_local_read_capability not importable")

        result = compile_document_local_read_workflow(
            "show rows where Score is greater than 85"
        )
        step = result["steps"][0]
        assert "tool_call" not in step, "Guidance step must not have a tool_call"
