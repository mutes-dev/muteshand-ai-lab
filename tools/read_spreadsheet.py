INPUT_SPEC = {
    "path": "string"
}

import os
import sys

BASE_PATH = os.path.abspath("E:/MutesHand")

_DEFAULT_MAX_SHEETS = 3


def _is_extensionless_or_xlsx(path: str) -> bool:
    """Return True if path ends with .xlsx or resolver confirms XLSX content."""
    lower = path.lower()
    if lower.endswith(".xlsx"):
        return True
    # Only probe if there is truly no extension
    basename = os.path.basename(path)
    if "." in basename:
        return False
    _project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    if _project_root not in sys.path:
        sys.path.insert(0, _project_root)
    from system.orchestrator.capabilities.document_intake_resolver import is_extensionless_acceptable
    return is_extensionless_acceptable("read_spreadsheet", path)


_DEFAULT_MAX_ROWS_PER_SHEET = 100
_DEFAULT_MAX_COLUMNS = 50
_DEFAULT_MAX_CELL_CHARS = 500


def _truncate_cell(value, max_chars):
    """Truncate a cell value to max_chars with a controlled note."""
    text = str(value) if value is not None else ""
    if len(text) > max_chars:
        return text[:max_chars] + " [additional cell content omitted]"
    return text


def run(
    path,
    max_sheets=None,
    max_rows_per_sheet=None,
    max_columns=None,
    max_cell_chars=None,
):
    """
    Extract a bounded structured preview from a local XLSX workbook.

    Uses openpyxl in read_only=True, data_only=True mode.
    Does not execute formulas or macros.
    Returns structured dict for all cases.
    """
    _max_sheets = max_sheets if isinstance(max_sheets, int) and max_sheets > 0 else _DEFAULT_MAX_SHEETS
    _max_rows_per_sheet = (
        max_rows_per_sheet if isinstance(max_rows_per_sheet, int) and max_rows_per_sheet > 0 else _DEFAULT_MAX_ROWS_PER_SHEET
    )
    _max_columns = max_columns if isinstance(max_columns, int) and max_columns > 0 else _DEFAULT_MAX_COLUMNS
    _max_cell_chars = max_cell_chars if isinstance(max_cell_chars, int) and max_cell_chars > 0 else _DEFAULT_MAX_CELL_CHARS

    try:
        _project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
        if _project_root not in sys.path:
            sys.path.insert(0, _project_root)
        from system.security.path_validator import validate_path

        validation = validate_path(path, BASE_PATH)
        if validation.get("status") == "failure":
            return validation

        full_path = validation["resolved_path"]

        if not os.path.exists(full_path):
            return {"status": "failure", "reason": "file_not_found"}

        lower = full_path.lower()
        if lower.endswith(".xls") or lower.endswith(".xlsm"):
            return {"status": "failure", "reason": "unsupported_format"}
        if not _is_extensionless_or_xlsx(full_path):
            return {"status": "failure", "reason": "unsupported_format"}

        from openpyxl import load_workbook

        # openpyxl requires .xlsx extension on string paths; use file object for extensionless
        if full_path.lower().endswith(".xlsx"):
            wb_formula = load_workbook(full_path, read_only=True, data_only=False)
            wb_value = load_workbook(full_path, read_only=True, data_only=True)
            _file_handles = []
        else:
            f_formula = open(full_path, "rb")
            f_value = open(full_path, "rb")
            wb_formula = load_workbook(f_formula, read_only=True, data_only=False)
            wb_value = load_workbook(f_value, read_only=True, data_only=True)
            _file_handles = [f_formula, f_value]
        sheet_names = wb_value.sheetnames

        previewed_sheet_blocks = []
        sheets_omitted = False
        any_rows_omitted = False
        any_columns_omitted = False
        any_cells_truncated = False

        try:
            for sheet_index, sheet_name in enumerate(sheet_names):
                if sheet_index >= _max_sheets:
                    sheets_omitted = True
                    break

                ws_formula = wb_formula[sheet_name]
                ws_value = wb_value[sheet_name]

                headers = None
                data_rows = []
                max_column_count = 0
                total_rows = 0
                rows_omitted = False
                columns_omitted = False
                cells_truncated = False

                for f_row, v_row in zip(ws_formula.iter_rows(), ws_value.iter_rows()):
                    total_rows += 1
                    values = []
                    for f_cell, v_cell in zip(f_row, v_row):
                        f_val = f_cell.value
                        v_val = v_cell.value

                        if f_val is not None and isinstance(f_val, str) and f_val.startswith("="):
                            # Formula cell: read formula, show cached value if available
                            cached_str = str(v_val) if v_val is not None else "unavailable"
                            display = f"Formula: {f_val}; Cached value: {cached_str}"
                        else:
                            display = str(v_val) if v_val is not None else ""

                        truncated = _truncate_cell(display, _max_cell_chars)
                        if truncated != display:
                            cells_truncated = True
                        values.append(truncated)

                    max_column_count = max(max_column_count, len(values))

                    if total_rows == 1:
                        headers = []
                        for idx, h in enumerate(values):
                            h_str = str(h) if h is not None else ""
                            if h_str.strip() == "":
                                headers.append(f"Column {idx + 1}")
                            else:
                                headers.append(h_str)
                        continue

                    if len(data_rows) >= _max_rows_per_sheet:
                        rows_omitted = True
                        continue

                    truncated_row = []
                    for idx, val in enumerate(values):
                        if idx >= _max_columns:
                            columns_omitted = True
                            break
                        truncated_row.append(val)

                    data_rows.append(truncated_row)

                # Pad headers to max column count
                if headers:
                    headers = headers + [""] * (max_column_count - len(headers))
                column_count = max_column_count

                block_lines = []
                block_lines.append(f"Sheet: {sheet_name}")
                block_lines.append(f"Dimensions: {total_rows}x{column_count}")
                block_lines.append(f"Preview rows shown: {len(data_rows)}")
                block_lines.append(f"Columns detected: {column_count}")
                if headers:
                    block_lines.append(f"Headers detected: {' | '.join(headers)}")
                else:
                    block_lines.append("Headers detected: none")

                block_lines.append("")
                block_lines.append("Preview:")
                if headers:
                    header_line = "| Row | " + " | ".join(headers) + " |"
                    block_lines.append(header_line)
                    for idx, row in enumerate(data_rows):
                        padded = row + [""] * (len(headers) - len(row))
                        block_lines.append(f"| {idx + 1} | " + " | ".join(padded) + " |")
                else:
                    block_lines.append("| Row | Data |")
                    for idx, row in enumerate(data_rows):
                        block_lines.append(f"| {idx + 1} | " + " | ".join(row) + " |")

                if rows_omitted:
                    block_lines.append(f"Additional rows omitted due to preview limit (max {_max_rows_per_sheet}).")
                    any_rows_omitted = True
                if columns_omitted:
                    block_lines.append(f"Additional columns omitted due to preview limit (max {_max_columns}).")
                    any_columns_omitted = True
                if cells_truncated:
                    block_lines.append(f"Some cell values shortened to {_max_cell_chars} characters.")
                    any_cells_truncated = True
                if total_rows > len(data_rows) + 1:
                    block_lines.append("Additional rows may exist beyond the preview.")

                previewed_sheet_blocks.append("\n".join(block_lines))

        finally:
            wb_formula.close()
            wb_value.close()
            for fh in _file_handles:
                fh.close()

        lines = []
        lines.append(f"Workbook: {os.path.basename(full_path)}")
        lines.append("Format: XLSX")
        lines.append(f"Total sheets: {len(sheet_names)}")
        lines.append(f"Sheet names: {', '.join(sheet_names)}")
        lines.append(f"Sheets previewed: {len(previewed_sheet_blocks)} of {len(sheet_names)}")

        if sheets_omitted:
            lines.append(f"Additional sheets omitted due to preview limit (max {_max_sheets}).")

        lines.append("")
        lines.append("Formula handling:")
        lines.append("- Formulas are read but not executed.")
        lines.append("- Cached formula values are shown only if saved in the workbook.")
        lines.append('- Formula cells without cached values show "cached value unavailable".')

        lines.append("")
        lines.append("\n\n".join(previewed_sheet_blocks))

        lines.append("")
        lines.append("Limits:")
        lines.append(f"- Max sheets: {_max_sheets}")
        lines.append(f"- Max preview rows per sheet: {_max_rows_per_sheet}")
        lines.append(f"- Max columns: {_max_columns}")
        lines.append(f"- Max cell characters: {_max_cell_chars}")

        lines.append("")
        lines.append("Notes:")
        lines.append("- This is a bounded workbook preview, not full workbook analysis.")
        lines.append("- Charts, images, macros, pivots, and formulas are not executed or interpreted.")
        lines.append("- Additional sheets/rows/columns/cell content may be omitted when limits are reached.")

        result_text = "\n".join(lines)

        return {
            "status": "success",
            "result": result_text,
            "metadata": {
                "source_path": full_path,
                "format": "xlsx",
                "sheet_names": sheet_names,
                "previewed_sheets": len(previewed_sheet_blocks),
                "sheets_omitted": sheets_omitted,
                "max_rows_per_sheet": _max_rows_per_sheet,
                "max_columns": _max_columns,
                "rows_omitted": any_rows_omitted,
                "columns_omitted": any_columns_omitted,
                "cells_truncated": any_cells_truncated,
                "formula_note": "Formulas are read but not executed. Cached formula values are shown only if saved in the workbook.",
            },
        }

    except Exception:
        return {"status": "failure", "reason": "read_error"}
