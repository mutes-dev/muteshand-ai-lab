INPUT_SPEC = {
    "file_path": "string",
    "reference_type": "string",
    "sheet_name": "string",
    "has_header": "boolean",
    "header_row": "number",
    "row_number": "number",
    "cell_address": "string",
    "column_name": "string",
    "column_index": "number",
    "entity_column": "string",
    "max_rows_scan": "number",
    "max_columns_scan": "number",
    "max_cell_chars": "number",
}

import os
import re
import sys

BASE_PATH = os.path.abspath("E:/MutesHand")

_DEFAULT_MAX_ROWS_SCAN = 1000
_DEFAULT_MAX_COLUMNS_SCAN = 50
_DEFAULT_MAX_CELL_CHARS = 500

_REFERENCE_TYPES = {"row", "cell", "entity_from_row"}

# Auto-discovery sentinel for name-like columns (e.g., "third name").
_AUTO_NAME_LIKE_COLUMN = "__AUTO_NAME_LIKE__"

# Headers that suggest a name/person/entity column.
_NAME_LIKE_HEADER_RE = re.compile(
    r"\b(?:name|names|person|people|user|users|entity|entities|contact|customer|client|"
    r"full\s*name|first\s*name|last\s*name)\b",
    re.IGNORECASE,
)


def _find_unique_name_like_column(headers):
    """Return the unique name-like header, or None if ambiguous/absent."""
    matches = [h for h in headers if _NAME_LIKE_HEADER_RE.search(h)]
    if len(matches) == 1:
        return matches[0]
    return None


def _ensure_project_root_in_sys_path():
    _project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    if _project_root not in sys.path:
        sys.path.insert(0, _project_root)


def _validate_file_path(path):
    _ensure_project_root_in_sys_path()
    from system.security.path_validator import validate_path
    return validate_path(path, BASE_PATH)


def _file_type_from_path(path):
    lower = path.lower()
    if lower.endswith(".csv"):
        return "csv"
    if lower.endswith(".xlsx"):
        return "xlsx"
    if lower.endswith(".xls") or lower.endswith(".xlsm"):
        return "unsupported"
    return "unsupported"


def _normalize_has_header(value):
    """Normalize has_header to bool, accepting string flags from tool calls."""
    if isinstance(value, bool):
        return value
    if isinstance(value, str):
        lower = value.strip().lower()
        if lower in ("1", "true", "yes"):
            return True
        if lower in ("0", "false", "no", ""):
            return False
    return bool(value)


def _parse_cell_address(cell_address):
    """Parse an Excel-style cell address into (column_index, row_number), both 1-based."""
    match = re.match(r"^([A-Za-z]+)(\d+)$", str(cell_address))
    if not match:
        return None, None
    letters = match.group(1).upper()
    row_number = int(match.group(2))
    col_index = 0
    for ch in letters:
        col_index = col_index * 26 + (ord(ch) - ord("A") + 1)
    return col_index, row_number


def _resolve_row_values(
    file_path,
    file_type,
    sheet_name,
    has_header,
    header_row,
    target_row,
    max_columns,
    max_cell_chars,
    max_rows_scan,
):
    """
    Scan a table until the target row is found.

    Returns (headers, header_map, row_values, truncated, out_of_bounds, row_not_found).
    """
    from tools.preview_table_schema import (
        _build_headers,
        _csv_row_generator,
        _xlsx_row_generator,
    )

    if file_type == "csv":
        gen = _csv_row_generator(file_path, max_columns, max_cell_chars)
        active_sheet = None
    else:
        from openpyxl import load_workbook

        wb = load_workbook(file_path, read_only=True, data_only=True)
        try:
            if sheet_name is None:
                active_sheet = wb.active.title
            else:
                if sheet_name not in wb.sheetnames:
                    return [], {}, None, False, False, "missing_sheet"
                active_sheet = sheet_name
        finally:
            wb.close()
        gen = _xlsx_row_generator(file_path, active_sheet, max_columns, max_cell_chars)

    headers = []
    header_map = {}
    truncated = False
    rows_scanned = 0

    for row_number, values, row_truncated in gen:
        if row_number > max_rows_scan + header_row:
            return headers, header_map, None, truncated or row_truncated, True, "row_not_found"

        rows_scanned += 1
        if row_truncated:
            truncated = True

        if has_header and row_number == header_row:
            headers = _build_headers(values)
            header_map = {header: idx + 1 for idx, header in enumerate(headers)}
            # If the reference explicitly targets the header row, return it
            # as a resolvable row (e.g., cell C1 or row 1).
            if target_row == header_row:
                return headers, header_map, values, truncated or row_truncated, False, None
            continue

        if row_number < header_row:
            continue

        if row_number == target_row:
            return headers, header_map, values, truncated or row_truncated, False, None

    return headers, header_map, None, truncated, False, "row_not_found"


def _error_result(file_path, error_code, message, reference_type=None, bounds_applied=None):
    return {
        "status": "error",
        "error_code": error_code,
        "tool": "resolve_table_reference",
        "reference_type": reference_type,
        "file_path": file_path,
        "message": message,
        "bounds_applied": bounds_applied or {},
    }


def run(
    file_path,
    reference_type,
    sheet_name=None,
    has_header=True,
    header_row=1,
    row_number=None,
    cell_address=None,
    column_name=None,
    column_index=None,
    entity_column=None,
    max_rows_scan=None,
    max_columns_scan=None,
    max_cell_chars=None,
):
    """
    Resolve a deterministic reference against a local CSV or XLSX table.

    Supported reference_type values:
    - row: return the full row at row_number.
    - cell: return a single cell value via cell_address or row_number + column_name/column_index.
    - entity_from_row: return the value in entity_column at row_number.

    This is reference resolution only — no analysis, aggregation, or formula execution.
    """
    _max_rows_scan = max_rows_scan if isinstance(max_rows_scan, int) and max_rows_scan > 0 else _DEFAULT_MAX_ROWS_SCAN
    _max_columns_scan = max_columns_scan if isinstance(max_columns_scan, int) and max_columns_scan > 0 else _DEFAULT_MAX_COLUMNS_SCAN
    _max_cell_chars = max_cell_chars if isinstance(max_cell_chars, int) and max_cell_chars > 0 else _DEFAULT_MAX_CELL_CHARS

    if sheet_name == "":
        sheet_name = None

    _bounds = {
        "max_rows_scan": _max_rows_scan,
        "max_columns_scan": _max_columns_scan,
        "max_cell_chars": _max_cell_chars,
    }

    try:
        if reference_type not in _REFERENCE_TYPES:
            return _error_result(file_path, "unsupported_reference_type", f"Unsupported reference_type '{reference_type}'. Use row, cell, or entity_from_row.", bounds_applied=_bounds)

        validation = _validate_file_path(file_path)
        if validation.get("status") == "failure":
            return validation

        full_path = validation["resolved_path"]

        if not os.path.exists(full_path):
            return _error_result(file_path, "file_not_found", "File not found.", reference_type, bounds_applied=_bounds)

        file_type = _file_type_from_path(full_path)
        if file_type == "unsupported":
            return _error_result(file_path, "unsupported_format", "Unsupported file format. Only .csv and .xlsx are supported.", reference_type, bounds_applied=_bounds)

        has_header = _normalize_has_header(has_header)
        if not isinstance(header_row, int) or header_row < 1:
            header_row = 1

        # Determine the target row and column for each reference type
        target_row = None
        target_col_index = None
        target_col_name = None
        target_cell_address = None

        if reference_type == "row":
            if not isinstance(row_number, int):
                return _error_result(file_path, "missing_row_number", "row_number is required for reference_type='row'.", reference_type, bounds_applied=_bounds)
            target_row = row_number

        elif reference_type == "cell":
            if cell_address:
                target_col_index, target_row = _parse_cell_address(cell_address)
                if target_col_index is None or target_row is None:
                    return _error_result(file_path, "invalid_cell_address", f"Invalid cell_address '{cell_address}'.", reference_type, bounds_applied=_bounds)
                target_cell_address = cell_address.upper()
            else:
                if not isinstance(row_number, int):
                    return _error_result(file_path, "missing_row_number", "row_number or cell_address is required for reference_type='cell'.", reference_type, bounds_applied=_bounds)
                target_row = row_number
                if column_name:
                    target_col_name = column_name
                elif isinstance(column_index, int):
                    target_col_index = column_index
                else:
                    return _error_result(file_path, "missing_column", "column_name or column_index is required when cell_address is not provided.", reference_type, bounds_applied=_bounds)

        elif reference_type == "entity_from_row":
            if not isinstance(row_number, int):
                return _error_result(file_path, "missing_row_number", "row_number is required for reference_type='entity_from_row'.", reference_type, bounds_applied=_bounds)
            if not entity_column:
                return _error_result(file_path, "missing_entity_column", "entity_column is required for reference_type='entity_from_row'.", reference_type, bounds_applied=_bounds)
            target_row = row_number
            target_col_name = entity_column

        # Resolve the target row and header map
        headers, header_map, row_values, truncated, out_of_bounds, reason = _resolve_row_values(
            full_path,
            file_type,
            sheet_name,
            has_header,
            header_row,
            target_row,
            _max_columns_scan,
            _max_cell_chars,
            _max_rows_scan,
        )

        if reason == "missing_sheet":
            return _error_result(file_path, "missing_sheet", f"Sheet '{sheet_name}' not found in workbook.", reference_type, bounds_applied=_bounds)
        if out_of_bounds or reason == "row_not_found":
            return _error_result(file_path, "row_not_found", f"Row {target_row} not found within scan bounds (max_rows_scan={_max_rows_scan}).", reference_type, bounds_applied=_bounds)
        if row_values is None:
            return _error_result(file_path, "row_not_found", f"Row {target_row} not found.", reference_type, bounds_applied=_bounds)

        # Resolve column index from header name if needed (case-insensitive)
        if target_col_name:
            if target_col_name == _AUTO_NAME_LIKE_COLUMN:
                auto_col = _find_unique_name_like_column(headers)
                if auto_col is None:
                    return _error_result(
                        file_path,
                        "ambiguous_name_like_column",
                        "Cannot determine a unique name-like column; please specify the column name explicitly.",
                        reference_type,
                        bounds_applied=_bounds,
                    )
                target_col_name = auto_col
            header_map_lower = {h.lower(): idx for h, idx in header_map.items()}
            target_col_name_lower = target_col_name.lower()
            if target_col_name_lower not in header_map_lower:
                return _error_result(file_path, "missing_column", f"Column '{target_col_name}' not found in headers.", reference_type, bounds_applied=_bounds)
            target_col_index = header_map_lower[target_col_name_lower]
            # Preserve the original header casing for the output metadata.
            target_col_name = next(
                (h for h, idx in header_map.items() if idx == target_col_index),
                target_col_name,
            )

        # Validate the resolved column index
        if target_col_index is not None:
            if target_col_index < 1 or target_col_index > len(row_values):
                return _error_result(file_path, "column_index_out_of_range", f"Column index {target_col_index} is out of range for row {target_row}.", reference_type, bounds_applied=_bounds)
            value = row_values[target_col_index - 1]
            if not target_col_name and headers and target_col_index <= len(headers):
                target_col_name = headers[target_col_index - 1]
        else:
            value = None

        # Build data_ref
        data_ref = {
            "file_path": file_path,
            "file_type": file_type,
            "sheet_name": sheet_name if file_type == "xlsx" else None,
            "row_number": target_row,
            "column_name": target_col_name,
            "column_index": target_col_index,
            "cell_address": target_cell_address,
        }

        warnings = []
        if truncated:
            warnings.append(f"Some cells were truncated to {_max_cell_chars} characters.")
        if out_of_bounds:
            warnings.append(f"Scan limited to max_rows_scan={_max_rows_scan}.")

        if file_type == "csv":
            data_ref.pop("sheet_name", None)

        payload = {
            "tool": "resolve_table_reference",
            "reference_type": reference_type,
            "file_path": file_path,
            "file_type": file_type,
            "sheet_name": sheet_name if file_type == "xlsx" else None,
            "row_number": target_row if reference_type in ("row", "cell", "entity_from_row") else None,
            "column_name": target_col_name,
            "column_index": target_col_index,
            "cell_address": target_cell_address,
            "value": value if reference_type in ("cell", "entity_from_row") else None,
            "row": row_values if reference_type == "row" else None,
            "entity": value if reference_type == "entity_from_row" else None,
            "header_map": header_map,
            "data_ref": data_ref,
            "bounds_applied": {
                "max_rows_scan": _max_rows_scan,
                "max_columns_scan": _max_columns_scan,
                "max_cell_chars": _max_cell_chars,
            },
            "warnings": warnings,
            "message": "Reference resolved successfully.",
        }

        return {"status": "success", **payload, "result": payload}

    except Exception as e:
        return _error_result(file_path, "read_error", f"Unexpected error resolving reference: {str(e)}", reference_type, bounds_applied=_bounds)
