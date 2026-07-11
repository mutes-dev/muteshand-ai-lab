INPUT_SPEC = {
    "file_path": "string",
    "sheet_name": "string",
    "has_header": "boolean",
    "header_row": "number",
    "max_rows": "number",
    "max_columns": "number",
    "max_cell_chars": "number",
}

import csv
import os
import re
import sys

BASE_PATH = os.path.abspath("E:/MutesHand")

_DEFAULT_MAX_ROWS = 100
_DEFAULT_MAX_COLUMNS = 50
_DEFAULT_MAX_CELL_CHARS = 500


def _ensure_project_root_in_sys_path():
    _project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    if _project_root not in sys.path:
        sys.path.insert(0, _project_root)


def _validate_file_path(path):
    _ensure_project_root_in_sys_path()
    from system.security.path_validator import validate_path
    return validate_path(path, BASE_PATH)


def _truncate_cell(value, max_chars):
    """Truncate a cell value to max_chars with a controlled note."""
    text = str(value) if value is not None else ""
    if len(text) > max_chars:
        return text[:max_chars] + " [additional cell content omitted]"
    return text


def _detect_csv_delimiter(sample):
    """Detect delimiter using csv.Sniffer with comma fallback."""
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t")
        return dialect.delimiter
    except Exception:
        return ","


def _normalize_has_header(value):
    """Normalize has_header to bool, accepting string flags from tool calls."""
    if isinstance(value, bool):
        return value
    if isinstance(value, str):
        lower = value.strip().lower()
        if lower in ("1", "true", "yes"):
            return True
        if lower in ("0", "false", "no", ""):
            return False
    return bool(value)


def _build_headers(raw_values):
    """Build a clean header list from raw first-row values."""
    headers = []
    for idx, h in enumerate(raw_values):
        h_str = str(h) if h is not None else ""
        if h_str.strip() == "":
            headers.append(f"Column {idx + 1}")
        else:
            headers.append(h_str)
    return headers


def _csv_row_generator(file_path, max_columns, max_cell_chars):
    """Yield (row_number, values, truncated) for each row in a CSV file."""
    with open(file_path, "r", encoding="utf-8") as f:
        sample = f.read(4096)
        f.seek(0)
        delimiter = _detect_csv_delimiter(sample)
        reader = csv.reader(f, delimiter=delimiter)
        for row_number, row in enumerate(reader, start=1):
            values = []
            truncated = False
            for idx, cell in enumerate(row):
                if idx >= max_columns:
                    truncated = True
                    break
                values.append(_truncate_cell(cell, max_cell_chars))
            yield row_number, values, truncated


def _xlsx_row_generator(file_path, sheet_name, max_columns, max_cell_chars):
    """Yield (row_number, values, truncated) for each row in an XLSX sheet."""
    from openpyxl import load_workbook

    wb = load_workbook(file_path, read_only=True, data_only=True)
    try:
        if sheet_name is None:
            ws = wb.active
        else:
            ws = wb[sheet_name]
        for row_number, row in enumerate(ws.iter_rows(), start=1):
            values = []
            truncated = False
            for idx, cell in enumerate(row):
                if idx >= max_columns:
                    truncated = True
                    break
                v = cell.value
                text = str(v) if v is not None else ""
                values.append(_truncate_cell(text, max_cell_chars))
            yield row_number, values, truncated
    finally:
        wb.close()


def _file_type_from_path(path):
    lower = path.lower()
    if lower.endswith(".csv"):
        return "csv"
    if lower.endswith(".xlsx"):
        return "xlsx"
    if lower.endswith(".xls") or lower.endswith(".xlsm"):
        return "unsupported"
    return "unsupported"


def _scan_table(
    file_path,
    file_type,
    sheet_name,
    has_header,
    header_row,
    max_rows,
    max_columns,
    max_cell_chars,
):
    """Scan a table and return preview metadata."""
    if file_type == "csv":
        row_gen = _csv_row_generator(file_path, max_columns, max_cell_chars)
        active_sheet = None
        sheets = None
    else:
        from openpyxl import load_workbook

        wb = load_workbook(file_path, read_only=True, data_only=True)
        try:
            sheet_names = wb.sheetnames
            if sheet_name is None:
                active_sheet = wb.active.title
            else:
                active_sheet = sheet_name
            sheets = list(sheet_names)
        finally:
            wb.close()
        row_gen = _xlsx_row_generator(file_path, active_sheet, max_columns, max_cell_chars)

    headers = []
    header_map = {}
    sample_rows = []
    rows_omitted = False
    columns_omitted = False
    row_count_observed = 0
    data_rows_collected = 0

    start_data_row = header_row + 1 if has_header else header_row

    for row_number, values, truncated in row_gen:
        row_count_observed = row_number

        if truncated:
            columns_omitted = True

        if has_header and row_number == header_row:
            headers = _build_headers(values)
            header_map = {header: idx + 1 for idx, header in enumerate(headers)}
            continue

        if row_number < start_data_row:
            continue

        if data_rows_collected >= max_rows:
            rows_omitted = True
            continue

        sample_rows.append({"row_number": row_number, "values": values})
        data_rows_collected += 1

    column_count_observed = len(headers) if headers else 0
    if not headers and sample_rows:
        column_count_observed = max(len(row["values"]) for row in sample_rows)

    warnings = []
    if columns_omitted:
        warnings.append(f"Additional columns omitted due to preview limit (max {max_columns}).")
    if rows_omitted:
        warnings.append(f"Additional rows omitted due to preview limit (max {max_rows}).")

    return {
        "headers": headers,
        "header_map": header_map,
        "sample_rows": sample_rows,
        "row_count_observed": row_count_observed,
        "column_count_observed": column_count_observed,
        "rows_omitted": rows_omitted,
        "columns_omitted": columns_omitted,
        "active_sheet": active_sheet,
        "sheets": sheets,
        "warnings": warnings,
    }


def run(
    file_path,
    sheet_name=None,
    has_header=True,
    header_row=1,
    max_rows=None,
    max_columns=None,
    max_cell_chars=None,
):
    """
    Return deterministic schema/preview metadata for a local CSV or XLSX table.

    This is reference support only — no analysis, aggregation, or formula execution.
    """
    _max_rows = max_rows if isinstance(max_rows, int) and max_rows > 0 else _DEFAULT_MAX_ROWS
    _max_columns = max_columns if isinstance(max_columns, int) and max_columns > 0 else _DEFAULT_MAX_COLUMNS
    _max_cell_chars = max_cell_chars if isinstance(max_cell_chars, int) and max_cell_chars > 0 else _DEFAULT_MAX_CELL_CHARS

    if sheet_name == "":
        sheet_name = None

    try:
        validation = _validate_file_path(file_path)
        if validation.get("status") == "failure":
            return validation

        full_path = validation["resolved_path"]

        if not os.path.exists(full_path):
            return {
                "status": "error",
                "error_code": "file_not_found",
                "tool": "preview_table_schema",
                "file_path": file_path,
                "message": "File not found.",
            }

        file_type = _file_type_from_path(full_path)
        if file_type == "unsupported":
            return {
                "status": "error",
                "error_code": "unsupported_format",
                "tool": "preview_table_schema",
                "file_path": file_path,
                "message": "Unsupported file format. Only .csv and .xlsx are supported.",
            }

        has_header = _normalize_has_header(has_header)
        if not isinstance(header_row, int) or header_row < 1:
            header_row = 1

        scan = _scan_table(
            full_path,
            file_type,
            sheet_name,
            has_header,
            header_row,
            _max_rows,
            _max_columns,
            _max_cell_chars,
        )

        payload = {
            "tool": "preview_table_schema",
            "file_path": file_path,
            "file_type": file_type,
            "sheet_name": scan["active_sheet"],
            "sheets": scan["sheets"],
            "row_count_observed": scan["row_count_observed"],
            "column_count_observed": scan["column_count_observed"],
            "headers": scan["headers"],
            "header_map": scan["header_map"],
            "sample_rows": scan["sample_rows"],
            "rows_omitted": scan["rows_omitted"],
            "columns_omitted": scan["columns_omitted"],
            "bounds_applied": {
                "max_rows": _max_rows,
                "max_columns": _max_columns,
                "max_cell_chars": _max_cell_chars,
            },
            "warnings": scan["warnings"],
            "message": "Schema preview generated successfully.",
        }

        # Remove sheet keys for CSV
        if file_type == "csv":
            payload.pop("sheet_name", None)
            payload.pop("sheets", None)

        return {"status": "success", **payload, "result": payload}

    except Exception as e:
        return {
            "status": "error",
            "error_code": "read_error",
            "tool": "preview_table_schema",
            "file_path": file_path,
            "message": f"Unexpected error reading table: {str(e)}",
        }
