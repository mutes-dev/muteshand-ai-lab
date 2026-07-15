INPUT_SPEC = {
    "path": "string",
    "operation": "string",
    "target_column": "string",
    "sheet_name": "string",
    "associated_column": "string",
    "filter_op": "string",
    "filter_value": "string",
    "filter_value_to": "string",
}

# ---------------------------------------------------------------------------
# Compatibility marker — legacy flat positional interface
# ---------------------------------------------------------------------------
# run() and _analyze_table_impl() remain the COMPATIBILITY ADAPTER for existing
# capability-compiled workflows and all accepted F5A/F5B-1 calls.
# run_plan() is the new TableAnalysisPlanV1 execution entry point.
# Neither entry point bypasses system_entry or governance.
_LEGACY_INTERFACE_VERSION = "F5A-F5B1-compat"
_PLAN_INTERFACE_VERSION = "TableAnalysisPlanV1"

# Plan-mode reserved token carried in the legacy 'operation' argument.
# This keeps the existing 8-argument flat signature unchanged while allowing
# TableAnalysisPlanV1 to be lowered through system_entry as a string.
_PLAN_MODE_OPERATION_TOKEN = "__table_analysis_plan_v1__"

# Conservative payload bound for the serialized plan argument.
_MAX_PLAN_JSON_BYTES = 8192

import csv
import datetime
import json
import os
import re
import sys
from decimal import Decimal, InvalidOperation
from typing import Any

BASE_PATH = os.path.abspath("E:/MutesHand")

MAX_FILE_SIZE_BYTES = 10 * 1024 * 1024
MAX_DATA_ROWS = 10000
MAX_COLUMNS = 100
MAX_WORKSHEETS = 1
MAX_TIED_ROWS = 100
MAX_CELL_CHARS = 500
MAX_OFFENDING_CELLS = 50

_ALLOWED_OPERATIONS = {"count_rows", "max", "min", "sum", "average", "overview", "filter"}

MAX_FILTER_RESULT_ROWS = 1000
MAX_FILTER_VALUE_LEN = 500

_TEXT_FILTER_OPS = frozenset([
    "eq", "neq", "contains", "not_contains", "starts_with", "ends_with",
    "is_blank", "is_not_blank",
])
_NUMERIC_FILTER_OPS = frozenset([
    "eq", "neq", "gt", "gte", "lt", "lte", "between",
])
_ALL_FILTER_OPS = _TEXT_FILTER_OPS | _NUMERIC_FILTER_OPS

# Reuse name-like detection from resolve_table_reference to keep semantics identical.
_NAME_LIKE_HEADER_RE = re.compile(
    r"\b(?:name|names|person|people|user|users|entity|entities|contact|customer|client|"
    r"full\s*name|first\s*name|last\s*name)\b",
    re.IGNORECASE,
)


_NAME_LIKE_COLUMN_SENTINEL = "__AUTO_NAME_LIKE__"


# ---------------------------------------------------------------------------
# Numeric parsing / validation
# ---------------------------------------------------------------------------

_NUMERIC_STRING_RE = re.compile(r"^[+-]?(?:\d+(?:\.\d*)?|\.\d+)$")


def _is_numeric_string(text: str) -> bool:
    """Return True for plain integer/decimal strings without formatting."""
    return bool(_NUMERIC_STRING_RE.fullmatch(text.strip()))


def _decimal_to_serializable(value: Decimal) -> str:
    """Render Decimal as a plain decimal string without scientific notation."""
    return format(value, "f")


def _is_integral(value: Decimal) -> bool:
    """Return True if the Decimal is mathematically an integer."""
    return value == value.to_integral_value()


def _parse_numeric(value, file_type: str = "csv"):
    """
    Parse a table cell value into a Decimal.

    Returns:
        (Decimal, numeric_kind) on success, where numeric_kind is "integer" or "decimal".
        (None, reason) on failure, where reason is one of:
            "blank", "formula_cell_present", "unsupported_value_type",
            "non_numeric_value_present".
    """
    if value is None:
        return (None, "blank")

    if isinstance(value, bool):
        return (None, "non_numeric_value_present")

    if isinstance(value, (datetime.datetime, datetime.date, datetime.time)):
        return (None, "non_numeric_value_present")

    if isinstance(value, (int, float, Decimal)):
        if isinstance(value, float):
            try:
                decimal_value = Decimal(str(value))
            except (InvalidOperation, ValueError):
                return (None, "non_numeric_value_present")
        else:
            decimal_value = Decimal(value)
        numeric_kind = "integer" if _is_integral(decimal_value) else "decimal"
        return (decimal_value, numeric_kind)

    if isinstance(value, str):
        text = value.strip()
        if text == "":
            return (None, "blank")
        # Formula cell detection applies only to XLSX. openpyxl returns strings
        # starting with '=' when read_only=True, data_only=False.
        if file_type == "xlsx" and text.startswith("="):
            return (None, "formula_cell_present")
        if _is_numeric_string(text):
            try:
                decimal_value = Decimal(text)
            except InvalidOperation:
                return (None, "non_numeric_value_present")
            numeric_kind = "integer" if _is_integral(decimal_value) else "decimal"
            return (decimal_value, numeric_kind)
        return (None, "non_numeric_value_present")

    return (None, "non_numeric_value_present")


# ---------------------------------------------------------------------------
# Path / file helpers
# ---------------------------------------------------------------------------

def _ensure_project_root_in_sys_path():
    _project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    if _project_root not in sys.path:
        sys.path.insert(0, _project_root)


def _validate_file_path(path):
    _ensure_project_root_in_sys_path()
    from system.security.path_validator import validate_path
    return validate_path(path, BASE_PATH)


def _file_type_from_path(path: str) -> str:
    lower = path.lower()
    if lower.endswith(".csv"):
        return "csv"
    if lower.endswith(".xlsx"):
        return "xlsx"
    if lower.endswith(".xls") or lower.endswith(".xlsm"):
        return "unsupported_legacy"
    return "unsupported"


def _truncate_text(value, max_chars: int = MAX_CELL_CHARS) -> str:
    text = str(value) if value is not None else ""
    if len(text) > max_chars:
        return text[:max_chars] + " [additional cell content omitted]"
    return text


# ---------------------------------------------------------------------------
# Column / header helpers
# ---------------------------------------------------------------------------

def _build_headers(raw_values):
    headers = []
    for idx, h in enumerate(raw_values):
        h_str = str(h).strip() if h is not None else ""
        if h_str == "":
            headers.append(f"Column {idx + 1}")
        else:
            headers.append(h_str)
    return headers


def _has_duplicate_normalized_headers(headers):
    seen = set()
    for h in headers:
        normalized = h.strip().lower()
        if normalized in seen:
            return True
        seen.add(normalized)
    return False


def _resolve_column(headers, column_name, for_what="target"):
    """
    Resolve a column name to a 1-based index using case-insensitive match.

    Returns (1-based index, canonical_name, reason) where reason is None on success
    or one of: "column_not_found", "duplicate_column_header".
    """
    if column_name is None or str(column_name).strip() == "":
        return (None, None, "column_not_found")

    if _has_duplicate_normalized_headers(headers):
        return (None, None, "duplicate_column_header")

    lookup = {h.strip().lower(): (idx, h) for idx, h in enumerate(headers)}
    normalized = str(column_name).strip().lower()
    if normalized in lookup:
        idx, canonical = lookup[normalized]
        return (idx + 1, canonical, None)
    return (None, None, "column_not_found")


def _resolve_auto_name_like_column(headers):
    """Return the unique name-like header, or None if ambiguous/absent."""
    matches = [h for h in headers if _NAME_LIKE_HEADER_RE.search(h)]
    if len(matches) == 1:
        return matches[0]
    return None


def _column_letter(index: int) -> str:
    from openpyxl.utils import get_column_letter
    return get_column_letter(index)


# ---------------------------------------------------------------------------
# Table scanning
# ---------------------------------------------------------------------------

def _detect_csv_delimiter(sample: str) -> str:
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t")
        return dialect.delimiter
    except Exception:
        return ","


def _scan_csv(full_path: str):
    with open(full_path, "r", encoding="utf-8") as f:
        sample = f.read(4096)
        f.seek(0)
        delimiter = _detect_csv_delimiter(sample)
        reader = csv.reader(f, delimiter=delimiter)

        headers = None
        data_rows = []
        total_rows = 0

        for i, row in enumerate(reader):
            total_rows += 1
            if i == 0:
                headers = _build_headers(row)
                if len(headers) > MAX_COLUMNS:
                    return (None, None, "unsupported", "analysis_bounds_exceeded")
                continue
            if len(data_rows) >= MAX_DATA_ROWS:
                return (None, None, "unsupported", "analysis_bounds_exceeded")
            data_rows.append(row)

    return (headers, data_rows, None, None)


def _scan_xlsx(full_path: str, sheet_name=None):
    from openpyxl import load_workbook

    wb = load_workbook(full_path, read_only=True, data_only=False)
    try:
        sheet_names = wb.sheetnames
        if sheet_name is None:
            if len(sheet_names) != 1:
                return (None, None, "ambiguous", "multiple_sheets_require_selection")
            active_sheet = sheet_names[0]
        else:
            normalized_request = str(sheet_name).strip().lower()
            matches = [(s, s.strip().lower()) for s in sheet_names]
            candidates = [s for s, norm in matches if norm == normalized_request]
            if len(candidates) == 0:
                return (None, None, "not_found", "sheet_not_found")
            if len(candidates) > 1:
                return (None, None, "ambiguous", "sheet_name_ambiguous")
            active_sheet = candidates[0]

        ws = wb[active_sheet]
        headers = None
        data_rows = []
        for row in ws.iter_rows():
            values = [cell.value for cell in row]
            if headers is None:
                headers = _build_headers(values)
                if len(headers) > MAX_COLUMNS:
                    return (None, None, "unsupported", "analysis_bounds_exceeded")
                continue
            if len(data_rows) >= MAX_DATA_ROWS:
                return (None, None, "unsupported", "analysis_bounds_exceeded")
            data_rows.append(values)

        return (headers, data_rows, None, active_sheet)
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# Row-result serialization (F5B common foundation)
# ---------------------------------------------------------------------------

def _column_letter_safe(index: int) -> str:
    """Return spreadsheet column letter for 1-based index; fallback on ImportError."""
    try:
        from openpyxl.utils import get_column_letter
        return get_column_letter(index)
    except Exception:
        return str(index)


def _serialize_row(
    headers: list,
    row: list,
    row_number: int,
    file_type: str,
) -> dict:
    """
    Serialize a single data row into the stable F5B row-result shape.

    row_number is 1-based data-row index (header = row 1, first data row = 2).
    Returns a dict with keys: row_number, row_ref, cells.
    """
    cells = []
    for col_idx, col_name in enumerate(headers):
        raw = row[col_idx] if col_idx < len(row) else None
        if isinstance(raw, bool):
            display = str(raw)
        elif raw is None:
            display = ""
        elif isinstance(raw, str) and file_type == "xlsx" and raw.startswith("="):
            display = raw  # formula string — never executed, preserved as text
        else:
            display = _truncate_text(raw)
        cells.append({
            "column_name": col_name,
            "column_index": col_idx + 1,
            "column_ref": f"column:{col_name}",
            "cell_ref": f"{_column_letter_safe(col_idx + 1)}{row_number + 1}",
            "value": display,
        })
    return {
        "row_number": row_number,
        "row_ref": f"row:{row_number}",
        "cells": cells,
    }


# ---------------------------------------------------------------------------
# Filter computation (F5B-1)
# ---------------------------------------------------------------------------

def _classify_column_for_filter(headers, data_rows, col_idx, file_type):
    """
    Classify a column as numeric, text, mixed, empty, or formula-containing.

    Returns one of: 'numeric', 'text', 'mixed', 'empty', 'formula'.
    """
    numeric_count = 0
    text_count = 0
    formula_count = 0
    blank_count = 0

    for row in data_rows:
        raw = row[col_idx] if col_idx < len(row) else None
        if raw is None:
            blank_count += 1
            continue
        if isinstance(raw, bool):
            text_count += 1
            continue
        if isinstance(raw, (int, float, type(None).__class__)):
            pass
        if isinstance(raw, str):
            text = raw.strip()
            if text == "":
                blank_count += 1
                continue
            if file_type == "xlsx" and text.startswith("="):
                formula_count += 1
                continue
        parsed, kind = _parse_numeric(raw, file_type=file_type)
        if parsed is not None:
            numeric_count += 1
        elif kind == "blank":
            blank_count += 1
        elif kind == "formula_cell_present":
            formula_count += 1
        elif kind == "non_numeric_value_present":
            text_count += 1
        else:
            text_count += 1

    if formula_count > 0:
        return "formula"
    total_nonblank = numeric_count + text_count
    if total_nonblank == 0:
        return "empty"
    if numeric_count > 0 and text_count > 0:
        return "mixed"
    if numeric_count > 0:
        return "numeric"
    return "text"


def _is_blank_cell(raw_value) -> bool:
    """Return True if value is blank per F5B-1 semantics (None / empty / whitespace-only)."""
    if raw_value is None:
        return True
    if isinstance(raw_value, str) and raw_value.strip() == "":
        return True
    return False


def _apply_text_filter(raw_value, filter_op: str, filter_value: str) -> bool:
    """
    Apply a text filter operator to a single cell value.

    Uses Unicode casefold for case-insensitive matching.
    Trims surrounding whitespace on both sides before comparison.
    Preserves original value in row output (comparison only).
    """
    if filter_op == "is_blank":
        return _is_blank_cell(raw_value)
    if filter_op == "is_not_blank":
        return not _is_blank_cell(raw_value)

    cell_str = (str(raw_value) if raw_value is not None else "").strip()
    needle = filter_value.strip()
    cell_cf = cell_str.casefold()
    needle_cf = needle.casefold()

    if filter_op == "eq":
        return cell_cf == needle_cf
    if filter_op == "neq":
        return cell_cf != needle_cf
    if filter_op == "contains":
        return needle_cf in cell_cf
    if filter_op == "not_contains":
        return needle_cf not in cell_cf
    if filter_op == "starts_with":
        return cell_cf.startswith(needle_cf)
    if filter_op == "ends_with":
        return cell_cf.endswith(needle_cf)
    return False


def _apply_numeric_filter(
    parsed_value,  # Decimal
    filter_op: str,
    filter_parsed,  # Decimal
    filter_parsed_to=None,  # Decimal | None (for between)
) -> bool:
    if filter_op == "eq":
        return parsed_value == filter_parsed
    if filter_op == "neq":
        return parsed_value != filter_parsed
    if filter_op == "gt":
        return parsed_value > filter_parsed
    if filter_op == "gte":
        return parsed_value >= filter_parsed
    if filter_op == "lt":
        return parsed_value < filter_parsed
    if filter_op == "lte":
        return parsed_value <= filter_parsed
    if filter_op == "between":
        return filter_parsed <= parsed_value <= filter_parsed_to
    return False


def _build_filter_answer_text(
    column_name: str,
    filter_op: str,
    filter_value: str,
    filter_value_to: str,
    matched_row_count: int,
    returned_row_count: int,
    truncated: bool,
    headers: list,
    matched_rows: list,
) -> str:
    """Generate a concise deterministic filter answer_text."""
    pred_desc = _describe_predicate(column_name, filter_op, filter_value, filter_value_to)

    if matched_row_count == 0:
        return f"No rows match {pred_desc}."

    if truncated:
        return (
            f"{matched_row_count:,} rows match {pred_desc}. "
            f"Showing the first {returned_row_count:,} rows in Details / Evidence."
        )

    # Small complete result: try to include name-like column values inline.
    name_col = _resolve_auto_name_like_column(headers)
    if name_col is not None and matched_row_count <= 10:
        name_idx = next(
            (i for i, h in enumerate(headers) if h.strip().lower() == name_col.strip().lower()),
            None,
        )
        if name_idx is not None:
            names = []
            for row_dict in matched_rows:
                cells = row_dict.get("cells", [])
                val = next(
                    (c["value"] for c in cells if c["column_index"] == name_idx + 1),
                    None,
                )
                if val and str(val).strip():
                    names.append(str(val).strip())
            if names:
                return (
                    f"{matched_row_count} row{'s' if matched_row_count != 1 else ''} match "
                    f"{pred_desc}: {', '.join(names)}."
                )

    if matched_row_count == 1:
        return f"1 row matches {pred_desc}. See Details / Evidence for the matched row."
    return (
        f"{matched_row_count} rows match {pred_desc}. "
        f"See Details / Evidence for the matched rows."
    )


def _describe_predicate(
    column_name: str, filter_op: str, filter_value: str, filter_value_to: str
) -> str:
    """Produce a human-readable predicate description for answer_text."""
    labels = {
        "eq": "equals",
        "neq": "does not equal",
        "contains": "contains",
        "not_contains": "does not contain",
        "starts_with": "starts with",
        "ends_with": "ends with",
        "is_blank": "is blank",
        "is_not_blank": "is not blank",
        "gt": "is greater than",
        "gte": "is greater than or equal to",
        "lt": "is less than",
        "lte": "is less than or equal to",
        "between": "is between",
    }
    op_label = labels.get(filter_op, filter_op)
    if filter_op in ("is_blank", "is_not_blank"):
        return f"{column_name} {op_label}"
    if filter_op == "between":
        return f"{column_name} {op_label} {filter_value} and {filter_value_to}"
    return f"{column_name} {op_label} {filter_value}"


def _filter_matching_indices(
    headers: list,
    data_rows: list,
    filter_column: str,
    filter_op: str,
    filter_value: str,
    filter_value_to: str,
    file_type: str,
) -> dict:
    """Return all 0-based row indices that match the predicate — no truncation.

    Used internally by _execute_plan_operations for intermediate AND filter
    passes where the 1000-row display cap must not apply.

    Returns {"status": "success", "matching_indices": [...]} or an error dict.
    """
    if filter_op not in _ALL_FILTER_OPS:
        return {"status": "unsupported", "status_reason": "unsupported_filter_op"}
    if filter_value and len(filter_value) > MAX_FILTER_VALUE_LEN:
        return {"status": "unsupported", "status_reason": "filter_value_too_long"}
    if filter_value_to and len(filter_value_to) > MAX_FILTER_VALUE_LEN:
        return {"status": "unsupported", "status_reason": "filter_value_too_long"}

    col_idx, canonical_col, reason = _resolve_column(headers, filter_column, for_what="filter")
    if reason is not None:
        return {
            "status": "not_found" if reason == "column_not_found" else "ambiguous",
            "status_reason": reason,
        }

    if filter_op in ("is_blank", "is_not_blank"):
        filter_value = ""
        filter_value_to = ""
    if filter_op == "between" and not (filter_value_to and filter_value_to.strip()):
        return {"status": "unsupported", "status_reason": "missing_filter_upper_value"}

    col_type = _classify_column_for_filter(headers, data_rows, col_idx - 1, file_type)
    if col_type == "formula":
        return {"status": "unsupported", "status_reason": "formula_cell_present"}
    if filter_op in _NUMERIC_FILTER_OPS and filter_op not in ("eq", "neq"):
        if col_type not in ("numeric",):
            return {"status": "unsupported", "status_reason": "column_type_mismatch"}
    if filter_op in ("contains", "not_contains", "starts_with", "ends_with"):
        if col_type == "numeric":
            return {"status": "unsupported", "status_reason": "operator_not_supported_for_numeric_column"}

    filter_parsed = None
    filter_parsed_to = None
    comparison_type = "text"

    if filter_op in ("gt", "gte", "lt", "lte", "between", "eq", "neq") and col_type == "numeric":
        comparison_type = "numeric"
        if filter_op not in ("is_blank", "is_not_blank") and filter_value is not None and filter_value.strip():
            try:
                filter_parsed = Decimal(filter_value.strip())
            except (InvalidOperation, ValueError):
                return {"status": "unsupported", "status_reason": "filter_value_not_numeric"}
        if filter_op == "between":
            try:
                filter_parsed_to = Decimal(filter_value_to.strip())
            except (InvalidOperation, ValueError):
                return {"status": "unsupported", "status_reason": "filter_upper_value_not_numeric"}
            if filter_parsed > filter_parsed_to:
                return {"status": "unsupported", "status_reason": "invalid_filter_range"}

    if col_type == "mixed":
        comparison_type = "text"
        filter_parsed = None

    matching_indices = []
    for idx, row in enumerate(data_rows):
        raw = row[col_idx - 1] if col_idx - 1 < len(row) else None
        if comparison_type == "numeric" and filter_parsed is not None:
            parsed_val, _ = _parse_numeric(raw, file_type=file_type)
            if parsed_val is None:
                continue
            matched = _apply_numeric_filter(parsed_val, filter_op, filter_parsed, filter_parsed_to)
        else:
            matched = _apply_text_filter(raw, filter_op, filter_value or "")
        if matched:
            matching_indices.append(idx)

    return {
        "status": "success",
        "matching_indices": matching_indices,
        "canonical_col": canonical_col,
        "col_type": col_type,
        "comparison_type": comparison_type,
    }


def _compute_filter(
    headers: list,
    data_rows: list,
    filter_column: str,
    filter_op: str,
    filter_value: str,
    filter_value_to: str,
    file_type: str,
) -> dict:
    """
    Execute bounded single-predicate filter and return result dict.

    Returns a result dict with status 'success' or an error-indicating dict
    with keys 'status' and 'status_reason'.
    """
    # --- Validate filter_op ---
    if filter_op not in _ALL_FILTER_OPS:
        return {"status": "unsupported", "status_reason": "unsupported_filter_op"}

    # --- Validate filter_value length ---
    if filter_value and len(filter_value) > MAX_FILTER_VALUE_LEN:
        return {"status": "unsupported", "status_reason": "filter_value_too_long"}
    if filter_value_to and len(filter_value_to) > MAX_FILTER_VALUE_LEN:
        return {"status": "unsupported", "status_reason": "filter_value_too_long"}

    # --- Resolve column ---
    col_idx, canonical_col, reason = _resolve_column(headers, filter_column, for_what="filter")
    if reason is not None:
        return {
            "status": "not_found" if reason == "column_not_found" else "ambiguous",
            "status_reason": reason,
        }

    # --- blank/not_blank operators require no value ---
    if filter_op in ("is_blank", "is_not_blank"):
        filter_value = ""
        filter_value_to = ""

    # --- between requires filter_value_to ---
    if filter_op == "between" and not (filter_value_to and filter_value_to.strip()):
        return {"status": "unsupported", "status_reason": "missing_filter_upper_value"}

    # --- Classify target column ---
    col_type = _classify_column_for_filter(headers, data_rows, col_idx - 1, file_type)

    if col_type == "formula":
        return {"status": "unsupported", "status_reason": "formula_cell_present"}

    # --- Numeric operators on non-numeric columns ---
    # Pure numeric range operators (gt/gte/lt/lte/between) require a fully numeric column.
    # Mixed columns cannot support deterministic numeric comparison.
    if filter_op in _NUMERIC_FILTER_OPS and filter_op not in ("eq", "neq"):
        if col_type not in ("numeric",):
            return {"status": "unsupported", "status_reason": "column_type_mismatch"}

    # --- Text-only operators on numeric column ---
    if filter_op in ("contains", "not_contains", "starts_with", "ends_with"):
        if col_type == "numeric":
            return {"status": "unsupported", "status_reason": "operator_not_supported_for_numeric_column"}

    # --- Parse numeric filter values where needed ---
    filter_parsed = None
    filter_parsed_to = None
    comparison_type = "text"

    use_numeric = (
        filter_op in _NUMERIC_FILTER_OPS
        and filter_op not in ("eq", "neq")
    ) or (
        filter_op in ("eq", "neq")
        and col_type in ("numeric",)
        and filter_value is not None
        and filter_value.strip() != ""
        and _is_numeric_string(filter_value.strip())
    )

    if filter_op in ("gt", "gte", "lt", "lte", "between", "eq", "neq") and col_type == "numeric":
        comparison_type = "numeric"
        if filter_op not in ("is_blank", "is_not_blank") and filter_value is not None and filter_value.strip():
            try:
                filter_parsed = Decimal(filter_value.strip())
            except (InvalidOperation, ValueError):
                return {"status": "unsupported", "status_reason": "filter_value_not_numeric"}
        if filter_op == "between":
            try:
                filter_parsed_to = Decimal(filter_value_to.strip())
            except (InvalidOperation, ValueError):
                return {"status": "unsupported", "status_reason": "filter_upper_value_not_numeric"}
            if filter_parsed > filter_parsed_to:
                return {"status": "unsupported", "status_reason": "invalid_filter_range"}

    # --- Scan rows ---
    matched_row_count = 0
    returned_rows = []
    warnings = []

    if col_type == "mixed":
        warnings.append("Target column has mixed value types; only deterministic string comparison applied.")
        comparison_type = "text"  # mixed always uses text comparison
        filter_parsed = None

    for data_row_idx, row in enumerate(data_rows):
        row_number = data_row_idx + 1  # 1-based data-row number (header = row 1, data rows start at 2)
        raw = row[col_idx - 1] if col_idx - 1 < len(row) else None

        # Apply predicate
        if comparison_type == "numeric" and filter_parsed is not None:
            parsed_val, _ = _parse_numeric(raw, file_type=file_type)
            if parsed_val is None:
                continue  # blank/non-numeric — excluded from numeric filter
            matched = _apply_numeric_filter(parsed_val, filter_op, filter_parsed, filter_parsed_to)
        else:
            # Text comparison (or is_blank/is_not_blank)
            matched = _apply_text_filter(raw, filter_op, filter_value or "")

        if matched:
            matched_row_count += 1
            if len(returned_rows) < MAX_FILTER_RESULT_ROWS:
                returned_rows.append(_serialize_row(headers, row, row_number, file_type))

    truncated = matched_row_count > MAX_FILTER_RESULT_ROWS
    returned_row_count = len(returned_rows)
    result_complete = not truncated

    limitations = []
    if truncated:
        limitations.append(
            f"Only the first {MAX_FILTER_RESULT_ROWS:,} of {matched_row_count:,} "
            f"matched rows are included in the result."
        )

    row_refs = [r["row_ref"] for r in returned_rows]
    # Cell refs: predicate column cells for matched rows
    cell_refs = [
        next(
            (c["cell_ref"] for c in r["cells"] if c["column_index"] == col_idx),
            "",
        )
        for r in returned_rows
    ]

    return {
        "status": "success",
        "column_ref": f"column:{canonical_col}",
        "canonical_col": canonical_col,
        "filter_op": filter_op,
        "filter_value": filter_value or "",
        "filter_value_to": filter_value_to or "",
        "comparison_type": comparison_type,
        "col_type": col_type,
        "rows_evaluated": len(data_rows),
        "matched_row_count": matched_row_count,
        "returned_row_count": returned_row_count,
        "result_complete": result_complete,
        "truncated": truncated,
        "returned_rows": returned_rows,
        "row_refs": row_refs,
        "cell_refs": cell_refs,
        "warnings": warnings,
        "limitations": limitations,
    }


def _build_filter_result(
    filter_result: dict,
    file_path: str,
    file_type: str,
    resolved_sheet,
    headers: list,
    filter_column_name: str,
    filter_op: str,
    filter_value: str,
    filter_value_to: str,
) -> dict:
    """Assemble the structured filter success payload."""
    matched_row_count = filter_result["matched_row_count"]
    returned_row_count = filter_result["returned_row_count"]
    truncated = filter_result["truncated"]
    canonical_col = filter_result["canonical_col"]

    answer_text = _build_filter_answer_text(
        canonical_col,
        filter_op,
        filter_value or "",
        filter_value_to or "",
        matched_row_count,
        returned_row_count,
        truncated,
        headers,
        filter_result["returned_rows"],
    )

    return {
        "status": "success",
        "status_reason": None,
        "operation": "filter",
        "input_file": file_path,
        "file_type": file_type,
        "sheet_name": resolved_sheet if file_type == "xlsx" else None,
        "predicate": {
            "column_name": canonical_col,
            "column_ref": filter_result["column_ref"],
            "operator": filter_op,
            "value": filter_value or "",
            "value_to": filter_value_to or None,
            "comparison_type": filter_result["comparison_type"],
        },
        "rows_evaluated": filter_result["rows_evaluated"],
        "matched_row_count": matched_row_count,
        "returned_row_count": returned_row_count,
        "result_complete": filter_result["result_complete"],
        "truncated": truncated,
        "rows": filter_result["returned_rows"],
        "column_refs": [filter_result["column_ref"]],
        "row_refs": filter_result["row_refs"],
        "cell_refs": filter_result["cell_refs"],
        "warnings": filter_result["warnings"],
        "limitations": filter_result["limitations"],
        "answer_text": answer_text,
    }


# ---------------------------------------------------------------------------
# Aggregate computation
# ---------------------------------------------------------------------------

def _row_has_any_nonblank(row):
    for cell in row:
        if cell is not None and str(cell).strip() != "":
            return True
    return False


def _compute_count_rows(data_rows):
    count = sum(1 for row in data_rows if _row_has_any_nonblank(row))
    return {
        "operation": "count_rows",
        "value": count,
        "value_kind": "integer",
        "rows_evaluated": len(data_rows),
        "numeric_cells": None,
        "blank_cells": None,
        "associated": None,
    }


def _compute_average_stats(numeric_values):
    """Return deterministic average metadata from a list of (Decimal, kind, row_index)."""
    from fractions import Fraction
    from decimal import localcontext, ROUND_HALF_EVEN

    exact_total = sum(Fraction(value) for value, _, _ in numeric_values)
    count = len(numeric_values)
    avg_fraction = exact_total / count

    def _terminates(frac):
        d = frac.denominator
        for p in (2, 5):
            while d % p == 0:
                d //= p
        return d == 1

    rounded = not _terminates(avg_fraction)

    if rounded:
        with localcontext() as ctx:
            ctx.prec = 12
            ctx.rounding = ROUND_HALF_EVEN
            display_avg = Decimal(avg_fraction.numerator) / Decimal(avg_fraction.denominator)
        precision = 12
    else:
        digits = max(len(str(abs(avg_fraction.numerator))), len(str(abs(avg_fraction.denominator)))) + 10
        with localcontext() as ctx:
            ctx.prec = max(28, digits)
            ctx.rounding = ROUND_HALF_EVEN
            display_avg = Decimal(avg_fraction.numerator) / Decimal(avg_fraction.denominator)
        precision = ctx.prec

    all_integral = all(kind == "integer" for _, kind, _ in numeric_values)
    value_kind = "integer" if _is_integral(display_avg) and all_integral else "decimal"

    return {
        "value": _decimal_to_serializable(display_avg),
        "value_kind": value_kind,
        "sum": _decimal_to_serializable(
            Decimal(exact_total.numerator) / Decimal(exact_total.denominator)
        ),
        "count": count,
        "precision": precision,
        "rounding_mode": "ROUND_HALF_EVEN",
        "rounded": rounded,
    }


def _collect_cell_refs(start_row, end_row, col_index):
    col_letter = _column_letter(col_index)
    return f"{col_letter}{start_row}:{col_letter}{end_row}"


def _cell_ref(col_index, row_number):
    return f"{_column_letter(col_index)}{row_number}"


def _compute_numeric_aggregate(
    headers,
    data_rows,
    operation,
    target_column,
    associated_column=None,
    header_row_number=1,
    file_type="csv",
):
    col_idx, canonical_target, reason = _resolve_column(headers, target_column, for_what="target")
    if reason is not None:
        return {"status": "not_found" if reason == "column_not_found" else "ambiguous", "status_reason": reason}

    # Determine requested associated column.
    requested_associated = None
    if associated_column == _NAME_LIKE_COLUMN_SENTINEL:
        auto_col = _resolve_auto_name_like_column(headers)
        if auto_col is None:
            return {"status": "ambiguous", "status_reason": "associated_column_ambiguous"}
        requested_associated = auto_col
    elif associated_column is not None and str(associated_column).strip() != "":
        requested_associated = str(associated_column).strip()

    associated_idx = None
    canonical_associated = None
    if requested_associated is not None:
        associated_idx, canonical_associated, assoc_reason = _resolve_column(
            headers, requested_associated, for_what="associated"
        )
        if assoc_reason is not None:
            return {
                "status": "not_found" if assoc_reason == "column_not_found" else "ambiguous",
                "status_reason": assoc_reason,
            }

    numeric_values = []
    blank_count = 0
    offending_cells = []
    formula_detected = []

    for row_index, row in enumerate(data_rows, start=header_row_number + 1):
        raw_value = row[col_idx - 1] if col_idx - 1 < len(row) else None
        parsed, kind = _parse_numeric(raw_value, file_type=file_type)
        if parsed is not None:
            numeric_values.append((parsed, kind, row_index))
        elif kind == "blank":
            blank_count += 1
        elif kind == "formula_cell_present":
            formula_detected.append((row_index, _truncate_text(raw_value)))
        else:
            offending_cells.append((row_index, _truncate_text(raw_value), kind))

    if formula_detected:
        return {
            "status": "unsupported",
            "status_reason": "formula_cell_present",
            "offending_cells": [
                {"row": row_idx, "cell": _cell_ref(col_idx, row_idx), "value_preview": preview}
                for row_idx, preview in formula_detected[:MAX_OFFENDING_CELLS]
            ],
        }

    if offending_cells:
        return {
            "status": "unsupported",
            "status_reason": "non_numeric_value_present",
            "offending_cells": [
                {"row": row_idx, "cell": _cell_ref(col_idx, row_idx), "value_preview": preview}
                for row_idx, preview, _ in offending_cells[:MAX_OFFENDING_CELLS]
            ],
        }

    if not numeric_values:
        return {"status": "not_found", "status_reason": "no_numeric_values"}

    # All parsed values have a kind recorded alongside them.
    all_integral = all(kind == "integer" for _, kind, _ in numeric_values)

    if operation == "sum":
        total = sum(value for value, _, _ in numeric_values)
        value_kind = "integer" if _is_integral(total) and all_integral else "decimal"
        return {
            "operation": operation,
            "value": _decimal_to_serializable(total),
            "value_kind": value_kind,
            "rows_evaluated": len(data_rows),
            "numeric_cells": len(numeric_values),
            "blank_cells": blank_count,
            "associated": None,
        }

    if operation == "average":
        avg_stats = _compute_average_stats(numeric_values)
        avg_stats.update({
            "operation": operation,
            "rows_evaluated": len(data_rows),
            "numeric_cells": len(numeric_values),
            "blank_cells": blank_count,
            "associated": None,
        })
        return avg_stats

    # max or min
    extreme_value = max((v for v, _, _ in numeric_values)) if operation == "max" else min((v for v, _, _ in numeric_values))
    tied_rows = [(row_idx, value, kind) for value, kind, row_idx in numeric_values if value == extreme_value]

    if len(tied_rows) > MAX_TIED_ROWS:
        return {"status": "unsupported", "status_reason": "tie_result_bounds_exceeded"}

    associated_entries = []
    if associated_idx is not None:
        for row_idx, _, _ in tied_rows:
            raw_assoc = data_rows[row_idx - header_row_number - 1][associated_idx - 1] if associated_idx - 1 < len(
                data_rows[row_idx - header_row_number - 1]
            ) else None
            associated_entries.append({
                "row": row_idx,
                "associated_column": canonical_associated,
                "associated_cell": _cell_ref(associated_idx, row_idx),
                "associated_value": _truncate_text(raw_assoc) if raw_assoc is not None else "",
            })

    value_kind = "integer" if all_integral and _is_integral(extreme_value) else "decimal"
    start_row = header_row_number + 1
    end_row = header_row_number + len(data_rows)

    return {
        "operation": operation,
        "value": _decimal_to_serializable(extreme_value),
        "value_kind": value_kind,
        "rows_evaluated": len(data_rows),
        "numeric_cells": len(numeric_values),
        "blank_cells": blank_count,
        "tied_row_count": len(tied_rows),
        "tied_rows": [row_idx for row_idx, _, _ in tied_rows],
        "associated": {
            "associated_column": canonical_associated,
            "associated_rows": associated_entries,
        } if associated_entries else None,
        "extreme_cells": [_cell_ref(col_idx, row_idx) for row_idx, _, _ in tied_rows],
    }


# ---------------------------------------------------------------------------
# Overview computation
# ---------------------------------------------------------------------------

def _compute_overview(headers, data_rows, file_type, file_path, sheet_name=None):
    """Return table-level and per-column bounded statistics without LLM insight."""
    data_row_count = sum(1 for row in data_rows if _row_has_any_nonblank(row))
    blank_row_count = len(data_rows) - data_row_count
    warnings = []

    if _has_duplicate_normalized_headers(headers):
        warnings.append("Duplicate headers detected; column references use the original names.")

    columns = []
    for col_idx, col_name in enumerate(headers):
        col_letter = _column_letter(col_idx + 1)
        numeric_values = []
        blank_count = 0
        formula_count = 0
        unsupported_count = 0
        text_count = 0
        distinct_text_values = set()

        for row_idx, row in enumerate(data_rows, start=2):
            raw_value = row[col_idx] if col_idx < len(row) else None
            parsed, kind = _parse_numeric(raw_value, file_type=file_type)
            if parsed is not None:
                numeric_values.append((parsed, kind, row_idx))
            elif kind == "blank":
                blank_count += 1
            elif kind == "formula_cell_present":
                formula_count += 1
            elif kind == "non_numeric_value_present":
                text_count += 1
                if raw_value is not None:
                    distinct_text_values.add(_truncate_text(raw_value))
            else:
                unsupported_count += 1

        total_cells = len(data_rows)
        nonblank_count = total_cells - blank_count
        col_info = {
            "column_name": col_name,
            "column_ref": col_letter,
            "blank_count": blank_count,
            "nonblank_count": nonblank_count,
        }

        if formula_count > 0:
            col_info["classification"] = "formula-containing"
            col_info["formula_count"] = formula_count
            col_info["warning"] = "Formula cells detected but were not executed."
        elif unsupported_count > 0 or (numeric_values and text_count > 0):
            col_info["classification"] = "mixed" if (numeric_values or text_count > 0) else "unsupported"
            col_info["warning"] = "Mixed or unsupported value types detected; numeric statistics omitted."
        elif numeric_values and text_count == 0:
            # Pure numeric column
            min_value = min((v for v, _, _ in numeric_values))
            max_value = max((v for v, _, _ in numeric_values))
            min_cells = [
                _cell_ref(col_idx + 1, row_idx)
                for v, _, row_idx in numeric_values
                if v == min_value
            ][:MAX_TIED_ROWS]
            max_cells = [
                _cell_ref(col_idx + 1, row_idx)
                for v, _, row_idx in numeric_values
                if v == max_value
            ][:MAX_TIED_ROWS]
            total = sum((v for v, _, _ in numeric_values))
            value_kind = "integer" if all(kind == "integer" for _, kind, _ in numeric_values) else "decimal"
            avg_stats = _compute_average_stats(numeric_values)
            end_row = 1 + len(data_rows)
            col_info["classification"] = "numeric"
            col_info["numeric_count"] = len(numeric_values)
            col_info["min"] = _decimal_to_serializable(min_value)
            col_info["max"] = _decimal_to_serializable(max_value)
            col_info["min_cells"] = min_cells
            col_info["max_cells"] = max_cells
            col_info["sum"] = _decimal_to_serializable(total)
            col_info["average"] = avg_stats["value"]
            col_info["average_value_kind"] = avg_stats["value_kind"]
            col_info["average_rounded"] = avg_stats["rounded"]
            col_info["average_precision"] = avg_stats["precision"]
            col_info["average_rounding_mode"] = avg_stats["rounding_mode"]
            col_info["contributing_range"] = _collect_cell_refs(2, end_row, col_idx + 1)
        elif text_count > 0:
            col_info["classification"] = "text"
            col_info["distinct_count"] = len(distinct_text_values)
        elif total_cells == 0 or blank_count == total_cells:
            col_info["classification"] = "empty"
        else:
            col_info["classification"] = "unknown"

        columns.append(col_info)

    return {
        "file_path": file_path,
        "file_type": file_type,
        "sheet_name": sheet_name,
        "data_row_count": data_row_count,
        "blank_row_count": blank_row_count,
        "column_count": len(headers),
        "column_names": list(headers),
        "columns": columns,
        "warnings": warnings,
    }


def _build_overview_answer_text(overview):
    """Generate a concise deterministic answer_text from the overview result."""
    data_row_count = overview["data_row_count"]
    column_count = overview["column_count"]
    text = f"The table contains {data_row_count} data rows and {column_count} columns."
    if overview["blank_row_count"] > 0:
        text += f" {overview['blank_row_count']} fully blank row(s) were excluded."

    first_numeric = next((c for c in overview["columns"] if c["classification"] == "numeric"), None)
    if first_numeric:
        qualifier = "approximately " if first_numeric.get("average_rounded") else ""
        text += (
            f" {first_numeric['column_name']} is numeric, ranging from {first_numeric['min']} to"
            f" {first_numeric['max']}, with a total of {first_numeric['sum']} and an average of"
            f" {qualifier}{first_numeric['average']}."
        )
    else:
        text += " No numeric columns were found."

    text += " See Details / Evidence for the complete column overview."
    return text


def _build_overview_result(file_path, file_type, sheet_name, headers, data_rows, overview):
    """Assemble the structured success result for the overview operation."""
    answer_text = _build_overview_answer_text(overview)
    result = {
        "status": "success",
        "status_reason": None,
        "operation": "overview",
        "input_file": file_path,
        "file_type": file_type,
        "sheet_name": sheet_name if file_type == "xlsx" else None,
        "data_row_count": overview["data_row_count"],
        "blank_row_count": overview["blank_row_count"],
        "column_count": overview["column_count"],
        "column_names": overview["column_names"],
        "columns": overview["columns"],
        "column_refs": [f"column:{c['column_name']}" for c in overview["columns"]],
        "row_refs": [],
        "cell_refs": [],
        "contributing_cell_ranges": [],
        "warnings": overview["warnings"],
        "limitations": [],
        "answer_text": answer_text,
    }
    for col in overview["columns"]:
        if col["classification"] == "numeric":
            result["cell_refs"].extend(col.get("min_cells", []))
            result["cell_refs"].extend(col.get("max_cells", []))
            if col.get("contributing_range"):
                result["contributing_cell_ranges"].append(col["contributing_range"])
    return result


# ---------------------------------------------------------------------------
# Result assembly
# ---------------------------------------------------------------------------

def _build_error_result(status, status_reason, file_type, file_path, sheet_name=None, detail=None):
    payload = {
        "status": status,
        "status_reason": status_reason,
        "operation": None,
        "input_file": file_path,
        "file_type": file_type,
        "sheet_name": sheet_name,
        "column_refs": [],
        "row_refs": [],
        "cell_refs": [],
        "warnings": [],
        "limitations": [],
    }
    if detail:
        payload["detail"] = detail
    return payload


def _build_success_result(operation, file_path, file_type, sheet_name, headers, target_column, computed, header_row=1):
    computed_operation = computed["operation"]
    value = computed["value"]
    value_kind = computed["value_kind"]
    rows_evaluated = computed["rows_evaluated"]
    numeric_cells = computed["numeric_cells"]
    blank_cells = computed["blank_cells"]

    column_refs = []
    if operation != "count_rows" and target_column:
        column_refs.append(f"column:{target_column}")

    associated = computed.get("associated")
    if associated and associated.get("associated_column"):
        column_refs.append(f"column:{associated['associated_column']}")

    row_refs = []
    cell_refs = []
    contributing_cell_ranges = []

    start_data_row = header_row + 1
    end_data_row = header_row + rows_evaluated

    if operation in ("max", "min"):
        extreme_cells = computed.get("extreme_cells", [])
        cell_refs.extend(extreme_cells)
        if associated:
            for entry in associated["associated_rows"]:
                cell_refs.append(entry["associated_cell"])
        row_refs.extend(computed.get("tied_rows", []))
        if target_column:
            # Locate canonical target column in headers for evidence range.
            target_idx, _, _ = _resolve_column(headers, target_column)
            if target_idx is not None:
                contributing_cell_ranges.append(
                    _collect_cell_refs(start_data_row, end_data_row, target_idx)
                )

    rounded = computed.get("rounded", False)
    answer_text = _build_answer_text(operation, target_column, value, value_kind, associated, rounded=rounded)

    limitations = []
    if computed.get("tied_row_count", 0) > 1:
        limitations.append(f"{computed['tied_row_count']} rows tied for the {operation}imum value.")
    if operation == "average" and rounded:
        limitations.append(
            f"Average rounded to {computed['precision']} significant digits using {computed['rounding_mode']}."
        )

    result = {
        "status": "success",
        "status_reason": None,
        "operation": computed_operation,
        "input_file": file_path,
        "file_type": file_type,
        "sheet_name": sheet_name if file_type == "xlsx" else None,
        "column_refs": column_refs,
        "row_refs": row_refs,
        "cell_refs": cell_refs,
        "contributing_cell_ranges": contributing_cell_ranges,
        "computed_value": value,
        "value_kind": value_kind,
        "rows_evaluated": rows_evaluated,
        "numeric_cells": numeric_cells,
        "blank_cells": blank_cells,
        "answer_text": answer_text,
        "associated": associated,
        "warnings": [],
        "limitations": limitations,
    }
    if operation == "average":
        result["computed_sum"] = computed.get("sum")
        result["computed_count"] = computed.get("count")
        result["precision"] = computed.get("precision")
        result["rounding_mode"] = computed.get("rounding_mode")
        result["rounded"] = computed.get("rounded", False)
    if operation in ("max", "min"):
        result["extreme_cells"] = computed.get("extreme_cells", [])
        result["tied_rows"] = computed.get("tied_rows", [])
        result["tied_row_count"] = computed.get("tied_row_count", 0)
    return result


def _build_answer_text(operation, target_column, value, value_kind, associated, rounded=False):
    if operation == "count_rows":
        return f"There are {value} data rows in the table."

    label = target_column or "value"
    if operation == "max":
        base = f"The highest {label} is {value}."
    elif operation == "min":
        base = f"The lowest {label} is {value}."
    elif operation == "sum":
        base = f"The sum of {label} is {value}."
    elif operation == "average":
        qualifier = "approximately " if rounded else ""
        base = f"The average of {label} is {qualifier}{value}."
    else:
        base = f"The {operation} of {label} is {value}."

    if associated and associated.get("associated_rows"):
        assoc_values = [entry.get("associated_value", "") for entry in associated["associated_rows"]]
        if len(assoc_values) == 1:
            base += f" This corresponds to {assoc_values[0]}."
        else:
            base += f" This corresponds to: {', '.join(assoc_values)}."

    return base


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

def _analyze_table_impl(
    path,
    operation,
    target_column=None,
    sheet_name=None,
    associated_column=None,
    filter_op=None,
    filter_value=None,
    filter_value_to=None,
):
    """
    Raw bounded deterministic aggregate analysis implementation.

    Returns the structured payload dict directly (status may be
    success/not_found/ambiguous/unsupported/blocked/failed).
    """
    if not isinstance(path, str) or not path.strip():
        return _build_error_result("unsupported", "missing_path", None, None)

    operation = (operation or "").strip().lower()
    if operation not in _ALLOWED_OPERATIONS:
        return _build_error_result("unsupported", "unsupported_operation", None, path)

    file_type = _file_type_from_path(path)
    if file_type == "unsupported_legacy":
        return _build_error_result("unsupported", "unsupported_file_type", file_type, path)
    if file_type == "unsupported":
        return _build_error_result("unsupported", "unsupported_file_type", file_type, path)

    validation = _validate_file_path(path)
    if validation.get("status") == "failure":
        return _build_error_result("blocked", validation.get("reason", "path_safety_blocked"), file_type, path)

    full_path = validation["resolved_path"]

    if not os.path.exists(full_path):
        return _build_error_result("not_found", "file_not_found", file_type, path)

    try:
        file_size = os.path.getsize(full_path)
    except OSError:
        return _build_error_result("failed", "file_stat_error", file_type, path)

    if file_size > MAX_FILE_SIZE_BYTES:
        return _build_error_result("unsupported", "analysis_bounds_exceeded", file_type, path)

    if file_type == "csv":
        headers, data_rows, status, reason = _scan_csv(full_path)
        resolved_sheet = None
    else:
        headers, data_rows, status, reason_or_sheet = _scan_xlsx(full_path, sheet_name)
        if status is not None:
            reason = reason_or_sheet
            resolved_sheet = None
        else:
            reason = None
            resolved_sheet = reason_or_sheet

    if status is not None:
        return _build_error_result(status, reason, file_type, path, resolved_sheet)

    if not headers:
        return _build_error_result("unsupported", "empty_table", file_type, path, resolved_sheet)

    if operation == "filter":
        # Duplicate headers: ambiguous target — decline (consistent with aggregate ops).
        if _has_duplicate_normalized_headers(headers):
            return _build_error_result("ambiguous", "duplicate_column_header", file_type, path, resolved_sheet)
        if not target_column or not str(target_column).strip():
            return _build_error_result("unsupported", "missing_target_column", file_type, path, resolved_sheet)
        filter_op_val = (filter_op or "").strip().lower()
        filter_value_val = (filter_value or "").strip()
        filter_value_to_val = (filter_value_to or "").strip()
        computed = _compute_filter(
            headers,
            data_rows,
            target_column,
            filter_op_val,
            filter_value_val,
            filter_value_to_val,
            file_type,
        )
        if computed.get("status") != "success":
            return _build_error_result(
                computed["status"],
                computed["status_reason"],
                file_type,
                path,
                resolved_sheet,
            )
        return _build_filter_result(
            computed,
            path,
            file_type,
            resolved_sheet,
            headers,
            target_column,
            filter_op_val,
            filter_value_val,
            filter_value_to_val,
        )

    if operation != "overview" and _has_duplicate_normalized_headers(headers):
        return _build_error_result("ambiguous", "duplicate_column_header", file_type, path, resolved_sheet)

    if operation == "overview":
        overview = _compute_overview(
            headers, data_rows, file_type, path, sheet_name=resolved_sheet
        )
        return _build_overview_result(
            path, file_type, resolved_sheet, headers, data_rows, overview
        )

    if operation == "count_rows":
        computed = _compute_count_rows(data_rows)
    else:
        if target_column is None or str(target_column).strip() == "":
            return _build_error_result("unsupported", "missing_target_column", file_type, path, resolved_sheet)
        computed = _compute_numeric_aggregate(
            headers,
            data_rows,
            operation,
            target_column,
            associated_column=associated_column,
            file_type=file_type,
        )
        if computed.get("status") in ("not_found", "ambiguous", "unsupported"):
            result_payload = _build_error_result(
                computed["status"],
                computed["status_reason"],
                file_type,
                path,
                resolved_sheet,
            )
            if "offending_cells" in computed:
                result_payload["offending_cells"] = computed["offending_cells"]
            return result_payload

    return _build_success_result(
        operation,
        path,
        file_type,
        resolved_sheet,
        headers,
        target_column if operation != "count_rows" else None,
        computed,
    )


def _build_controlled_domain_answer_text(status_reason: str, target_column: str | None) -> str:
    """
    Return a concise deterministic answer_text for controlled terminal domain outcomes.

    These outcomes arise after the file was opened and the table was inspected.
    They are NOT unexpected failures — they are deterministic classifications that
    the tool can report precisely to the user.
    """
    col = target_column.strip() if target_column and str(target_column).strip() else "the requested column"
    _MESSAGES = {
        "column_type_mismatch": (
            f"Numeric range comparisons (greater than, less than, between) cannot be applied "
            f"to the text column \"{col}\". Use a text operator such as contains, starts with, "
            f"or equals instead."
        ),
        "operator_not_supported_for_numeric_column": (
            f"Text search operators (contains, starts with, ends with) cannot be applied to "
            f"the numeric column \"{col}\". Use a numeric operator such as greater than or equals instead."
        ),
        "formula_cell_present": (
            f"The column \"{col}\" contains formula cells. "
            f"Formula-based filtering is not supported in the current bounded implementation."
        ),
        "filter_value_not_numeric": (
            f"The filter value for column \"{col}\" must be a number for the requested numeric operator."
        ),
        "filter_upper_value_not_numeric": (
            f"The upper bound value for the between filter on column \"{col}\" must be a number."
        ),
        "missing_filter_upper_value": (
            f"The between operator on column \"{col}\" requires both a lower and upper value."
        ),
        "unsupported_filter_op": (
            "The requested filter operator is not supported. Supported operators: "
            "eq, neq, contains, not_contains, starts_with, ends_with, "
            "is_blank, is_not_blank, gt, gte, lt, lte, between."
        ),
        "filter_value_too_long": (
            "The filter value exceeds the maximum allowed length."
        ),
        "column_not_found": (
            f"Column \"{col}\" was not found in the table. "
            f"Check the column name and re-submit."
        ),
        "missing_target_column": (
            "A target column name is required for this filter operation."
        ),
        "ambiguous_column_name": (
            f"Column \"{col}\" matches more than one column header in the table. "
            f"Use a more specific column name."
        ),
        "duplicate_column_header": (
            "The table contains duplicate column headers. "
            "Filtering is not supported when column headers are ambiguous."
        ),
        "file_not_found": (
            "The requested file was not found. Check the path and re-submit."
        ),
        "analysis_bounds_exceeded": (
            "The file exceeds the maximum supported size for bounded analysis."
        ),
        "unsupported_operation": (
            "The requested operation is not supported. Supported operations: "
            "count_rows, min, max, sum, average, overview, filter."
        ),
        "unsupported_file_type": (
            "Only CSV (.csv) and XLSX (.xlsx) files are supported."
        ),
        "empty_table": (
            "The table appears to contain no column headers or data rows."
        ),
    }
    return _MESSAGES.get(
        status_reason,
        f"The requested operation could not be completed: {status_reason}.",
    )


def _is_controlled_domain_outcome(payload: dict) -> bool:
    """
    Return True when the payload represents a deterministic terminal domain outcome
    that arose after valid tool execution (file opened and inspected, or path resolved
    to a deterministic classification).

    These outcomes must complete the workflow once rather than triggering retries.

    Outer failure status is preserved for:
    - blocked: path safety policy violations
    - failed: infrastructure errors (file_stat_error, scan parse failures)
    """
    status = payload.get("status")
    reason = payload.get("status_reason", "")

    if status == "blocked":
        return False

    if status == "failed":
        return False

    if status in ("unsupported", "not_found", "ambiguous"):
        return True

    return False


# ---------------------------------------------------------------------------
# Plan-mode adapter: TableAnalysisPlanV1 lowered through the legacy flat signature
# ---------------------------------------------------------------------------

def _serialize_plan(plan: dict) -> str:
    """Deterministic, bounded JSON serialization of a TableAnalysisPlanV1."""
    return json.dumps(plan, sort_keys=True, separators=(",", ":"))


def _deserialize_plan(serialized: str) -> dict | None:
    """Safely decode a plan payload.  No eval, no executable syntax."""
    if not isinstance(serialized, str):
        return None
    try:
        payload = json.loads(serialized)
    except Exception:
        return None
    if not isinstance(payload, dict):
        return None
    return payload


def _normalize_source_path(path: str) -> str:
    """Normalize a path for source-mismatch comparison."""
    return path.replace("\\", "/").strip()


def _run_plan_mode(
    path: str,
    serialized_plan: str,
    sheet_name=None,
    associated_column=None,
    filter_op=None,
    filter_value=None,
    filter_value_to=None,
) -> dict:
    """Decode, validate, and execute a TableAnalysisPlanV1 carried as a string.

    This is the production lowering path:
      system_entry -> analyze_table.run() -> _run_plan_mode() -> run_plan()

    Rejects malformed, oversized, or mismatched payloads before execution.
    Legacy positional arguments must be exactly the values derived from the plan;
    any conflict causes a hard failure so the serialized plan remains the only
    execution authority.
    """
    try:
        from system.orchestrator.structured_data.table_analysis_plan import (
            validate_plan,
            PLAN_VERSION,
        )
    except ImportError as exc:
        return {
            "status": "failure",
            "reason": f"plan_module_import_error:{exc}",
            "observation": None,
        }

    if not isinstance(serialized_plan, str) or not serialized_plan.strip():
        return {
            "status": "failure",
            "reason": "plan_payload_empty",
            "observation": None,
        }

    if len(serialized_plan.encode("utf-8")) > _MAX_PLAN_JSON_BYTES:
        return {
            "status": "failure",
            "reason": "plan_payload_oversized",
            "observation": {"max_bytes": _MAX_PLAN_JSON_BYTES},
        }

    plan = _deserialize_plan(serialized_plan)
    if plan is None:
        return {
            "status": "failure",
            "reason": "plan_payload_not_valid_json",
            "observation": None,
        }

    if plan.get("version") != PLAN_VERSION:
        return {
            "status": "failure",
            "reason": "plan_version_mismatch",
            "observation": {"expected": PLAN_VERSION, "received": plan.get("version")},
        }

    validation = validate_plan(plan)
    if validation["status"] != "success":
        return {
            "status": "failure",
            "reason": f"plan_validation_failed:{validation.get('reason')}",
            "observation": validation,
        }

    source = plan.get("source", {})
    plan_path = source.get("path", "")
    if _normalize_source_path(plan_path) != _normalize_source_path(path):
        return {
            "status": "failure",
            "reason": "plan_source_path_mismatch",
            "observation": {
                "explicit_path": path,
                "plan_path": plan_path,
            },
        }

    # Legacy positional values must match the plan-derived values used by the
    # capability to lower the tool_call. If a caller injects a conflicting
    # positional value, the plan is tampered with and execution is rejected.
    operations = plan.get("operations", [])
    first_op = operations[0] if operations and isinstance(operations[0], dict) else {}

    expected_positional = {
        "sheet_name": (source.get("sheet") or ""),
        "associated_column": (first_op.get("associated_column") or ""),
        "filter_op": (first_op.get("filter_op") or ""),
        "filter_value": (first_op.get("filter_value") or ""),
        "filter_value_to": (first_op.get("filter_value_to") or ""),
    }
    actual_positional = {
        "sheet_name": (sheet_name or ""),
        "associated_column": (associated_column or ""),
        "filter_op": (filter_op or ""),
        "filter_value": (filter_value or ""),
        "filter_value_to": (filter_value_to or ""),
    }

    def _normalize_positional(v: Any) -> str:
        if v is None:
            return ""
        if isinstance(v, str):
            return v.replace("\\", "/").strip()
        return str(v).replace("\\", "/").strip()

    for field in expected_positional:
        if _normalize_positional(expected_positional[field]) != _normalize_positional(actual_positional[field]):
            return {
                "status": "failure",
                "reason": f"plan_positional_conflict:{field}",
                "observation": {
                    "field": field,
                    "expected": expected_positional[field],
                    "actual": actual_positional[field],
                },
            }

    # All defensive checks passed; execute through the validated run_plan path.
    return run_plan(plan)


def run(
    path,
    operation,
    target_column=None,
    sheet_name=None,
    associated_column=None,
    filter_op=None,
    filter_value=None,
    filter_value_to=None,
):
    """
    Public entry point used by system_entry — LEGACY COMPATIBILITY ADAPTER.

    Preserves the F5A/F5B-1 flat positional interface exactly.
    Adds additive trust metadata to the result payload without changing
    the outer wrapper shape, execution authority, or governance behavior.

    Returns:
        - tool execution success: {"status": "success", "result": <structured payload>}
        - controlled domain outcome: {"status": "success", "result": <structured payload
              with inner status unsupported/not_found/ambiguous and answer_text>}
        - genuine failure: {"status": "failure", "reason": <status_reason>,
                            "observation": <structured payload>}

    Controlled domain outcomes (column_type_mismatch, formula_cell_present, etc.) are
    wrapped in outer success so governance completes the workflow in one attempt rather
    than retrying an identical deterministic result.

    Genuine failures (path safety blocked, file_stat_error, scan parse failure) retain
    outer failure status — these are infrastructure or policy errors, not domain results.
    """
    # F5R: explicit plan-execution mode.  The legacy 8-argument signature is reused:
    # operation == _PLAN_MODE_OPERATION_TOKEN and target_column carries the JSON plan.
    if operation == _PLAN_MODE_OPERATION_TOKEN:
        return _run_plan_mode(
            path,
            target_column,
            sheet_name=sheet_name,
            associated_column=associated_column,
            filter_op=filter_op,
            filter_value=filter_value,
            filter_value_to=filter_value_to,
        )

    payload = _analyze_table_impl(
        path,
        operation,
        target_column=target_column,
        sheet_name=sheet_name,
        associated_column=associated_column,
        filter_op=filter_op,
        filter_value=filter_value,
        filter_value_to=filter_value_to,
    )

    if payload.get("status") == "success":
        payload = _attach_legacy_trust_metadata(payload, operation, path)
        return {"status": "success", "result": payload}

    if _is_controlled_domain_outcome(payload):
        status_reason = payload.get("status_reason") or payload.get("status") or "unknown"
        if "answer_text" not in payload or not payload.get("answer_text"):
            payload = dict(payload)
            payload["answer_text"] = _build_controlled_domain_answer_text(
                status_reason, target_column
            )
        payload = _attach_controlled_outcome_trust_metadata(payload, status_reason, operation)
        return {"status": "success", "result": payload}

    return {
        "status": "failure",
        "reason": payload.get("status_reason") or payload.get("status"),
        "observation": payload,
    }


# ---------------------------------------------------------------------------
# Trust metadata attachment helpers
# ---------------------------------------------------------------------------

def _attach_legacy_trust_metadata(payload: dict, operation: str, path: str) -> dict:
    """Attach additive trust metadata to a legacy single-operation success result.

    trust_class=verified for deterministic complete results.
    Additive only — does not alter any existing field.
    """
    payload = dict(payload)
    op_id = f"op_{operation}"
    evidence_refs = (
        payload.get("cell_refs", [])
        + payload.get("row_refs", [])
        + payload.get("column_refs", [])
    )
    result_complete = payload.get("result_complete", True)
    limitations = payload.get("limitations", [])
    warnings = payload.get("warnings", [])
    payload["trust_metadata"] = {
        "trust_class": "verified",
        "verification_status": "verified",
        "plan_version": _PLAN_INTERFACE_VERSION,
        "plan_source_path": path,
        "requested_operations": [op_id],
        "executed_operations": [op_id],
        "omitted_operations": [],
        "operation_coverage_complete": True,
        "result_complete": result_complete,
        "evidence_refs": [str(r) for r in evidence_refs if r],
        "source_context_refs": [],
        "context_scope": "deterministic_full_scan",
        "context_complete": True,
        "advisory_disclaimer": None,
        "unsupported_reason": None,
        "ambiguity_reason": None,
        "clarification_needed": False,
        "limitations": limitations,
        "warnings": warnings,
        "learning_eligible": False,
        "operator_acceptance_status": "unreviewed",
    }
    return payload


def _attach_controlled_outcome_trust_metadata(
    payload: dict, status_reason: str, operation: str
) -> dict:
    """Attach trust metadata to controlled domain outcome payloads.

    unsupported domain outcomes → trust_class=unsupported
    ambiguous domain outcomes   → trust_class=ambiguous
    not_found outcomes          → trust_class=ambiguous (source/column unresolvable)
    """
    payload = dict(payload)
    outer_status = payload.get("status", "")
    if outer_status == "unsupported":
        trust_class = "unsupported"
        verification_status = "not_applicable"
        unsupported_reason = status_reason
        ambiguity_reason = None
        clarification_needed = False
    elif outer_status in ("ambiguous", "not_found"):
        trust_class = "ambiguous"
        verification_status = "not_applicable"
        unsupported_reason = None
        ambiguity_reason = status_reason
        clarification_needed = True
    else:
        trust_class = "unsupported"
        verification_status = "not_applicable"
        unsupported_reason = status_reason
        ambiguity_reason = None
        clarification_needed = False

    op_id = f"op_{operation}" if operation else "op_unknown"
    payload["trust_metadata"] = {
        "trust_class": trust_class,
        "verification_status": verification_status,
        "plan_version": _PLAN_INTERFACE_VERSION,
        "plan_source_path": payload.get("input_file"),
        "requested_operations": [op_id],
        "executed_operations": [],
        "omitted_operations": [op_id],
        "operation_coverage_complete": False,
        "result_complete": True,
        "evidence_refs": [],
        "source_context_refs": [],
        "context_scope": None,
        "context_complete": None,
        "advisory_disclaimer": None,
        "unsupported_reason": unsupported_reason,
        "ambiguity_reason": ambiguity_reason,
        "clarification_needed": clarification_needed,
        "limitations": payload.get("limitations", []),
        "warnings": payload.get("warnings", []),
        "learning_eligible": False,
        "operator_acceptance_status": "unreviewed",
    }
    return payload


# ---------------------------------------------------------------------------
# Plan execution — TableAnalysisPlanV1 entry point
# ---------------------------------------------------------------------------

def run_plan(plan: dict) -> dict:
    """Execute a validated TableAnalysisPlanV1.

    This is the new F5R entry point for composed multi-operation plans.
    It does NOT replace system_entry or governance.
    Callers must have already validated the plan via
    system.orchestrator.structured_data.table_analysis_plan.validate_plan().

    Returns the same outer shape as run():
        {"status": "success"|"failure", "result"|"reason"|"observation": ...}

    The result payload contains trust_metadata with full coverage fields.
    """
    try:
        from system.orchestrator.structured_data.table_analysis_plan import (
            validate_plan,
            validate_coverage,
            PLAN_VERSION,
            TRUST_CLASS_VERIFIED,
            TRUST_CLASS_UNSUPPORTED,
            TRUST_CLASS_AMBIGUOUS,
        )
    except ImportError as exc:
        return {"status": "failure", "reason": f"plan_module_import_error:{exc}"}

    validation = validate_plan(plan)
    if validation["status"] != "success":
        return {
            "status": "failure",
            "reason": f"plan_validation_failed:{validation.get('reason')}",
            "observation": validation,
        }

    source = plan["source"]
    path = source["path"]
    sheet_name = source.get("sheet")
    operations = plan["operations"]
    requested_ops = plan.get("requested_operations", [])

    executed_ids: list[str] = []
    payload = _execute_plan_operations(
        path=path,
        sheet_name=sheet_name,
        operations=operations,
        executed_ids=executed_ids,
    )

    if payload.get("status") == "plan_error":
        return {
            "status": "failure",
            "reason": payload.get("reason", "plan_execution_error"),
            "observation": payload,
        }

    coverage = validate_coverage(plan, executed_ids)
    trust_class = TRUST_CLASS_VERIFIED if coverage["operation_coverage_complete"] else TRUST_CLASS_UNSUPPORTED

    # F5R-FIX4: if the plan required an associated row/entity, verified depends on
    # the result actually containing associated evidence.
    if trust_class == TRUST_CLASS_VERIFIED:
        for op in plan.get("operations", []):
            if op.get("associated_column"):
                if (
                    payload.get("status") != "success"
                    or not payload.get("associated")
                    or not payload["associated"].get("associated_rows")
                ):
                    trust_class = TRUST_CLASS_UNSUPPORTED
                    break

    # The deterministic engine itself can produce controlled-domain ambiguity
    # (e.g., no unique name-like column). Preserve that classification.
    if payload.get("status") == "ambiguous":
        trust_class = TRUST_CLASS_AMBIGUOUS
    elif payload.get("status") == "unsupported":
        trust_class = TRUST_CLASS_UNSUPPORTED

    evidence_refs = (
        payload.get("cell_refs", [])
        + payload.get("row_refs", [])
        + payload.get("column_refs", [])
    )

    payload["trust_metadata"] = {
        "trust_class": trust_class,
        "verification_status": "verified" if trust_class == TRUST_CLASS_VERIFIED else "not_applicable",
        "plan_version": PLAN_VERSION,
        "plan_source_path": path,
        "requested_operations": coverage["requested_operations"],
        "executed_operations": coverage["executed_operations"],
        "omitted_operations": coverage["omitted_operations"],
        "operation_coverage_complete": coverage["operation_coverage_complete"],
        "result_complete": payload.get("result_complete", coverage["result_complete"]),
        "evidence_refs": [str(r) for r in evidence_refs if r],
        "source_context_refs": [],
        "context_scope": "deterministic_full_scan",
        "context_complete": True,
        "advisory_disclaimer": None,
        "unsupported_reason": None if trust_class != TRUST_CLASS_UNSUPPORTED else (payload.get("status_reason") or "partial_execution"),
        "ambiguity_reason": (payload.get("status_reason") or "associated_column_ambiguous") if trust_class == TRUST_CLASS_AMBIGUOUS else None,
        "clarification_needed": trust_class == TRUST_CLASS_AMBIGUOUS,
        "limitations": payload.get("limitations", []),
        "warnings": payload.get("warnings", []),
        "learning_eligible": False,
        "operator_acceptance_status": "unreviewed",
    }

    outer_status = payload.get("status", "success")
    if outer_status == "success":
        return {"status": "success", "result": payload}

    if _is_controlled_domain_outcome(payload):
        if "answer_text" not in payload or not payload.get("answer_text"):
            payload["answer_text"] = _build_controlled_domain_answer_text(
                payload.get("status_reason") or "", None
            )
        return {"status": "success", "result": payload}

    return {
        "status": "failure",
        "reason": payload.get("status_reason") or payload.get("status"),
        "observation": payload,
    }


def _execute_plan_operations(
    path: str,
    sheet_name,
    operations: list,
    executed_ids: list,
) -> dict:
    """Execute the ordered operation list from a TableAnalysisPlanV1.

    Supported composition for F5R:
    - Multiple filter operations (AND semantics — each narrows the row set)
    - One optional sort operation (stable, deterministic)
    - Single aggregate / overview (non-composed, falls back to legacy impl)

    For single-operation plans the legacy _analyze_table_impl path is used
    to preserve exact existing behavior and evidence shapes.
    """
    if not operations:
        return {"status": "plan_error", "reason": "empty_operations"}

    filter_ops = [op for op in operations if op.get("type") == "filter"]
    sort_ops = [op for op in operations if op.get("type") == "sort"]
    other_ops = [op for op in operations if op.get("type") not in ("filter", "sort")]

    # Single non-filter/sort operation: delegate to legacy path
    if len(filter_ops) == 0 and len(sort_ops) == 0 and len(other_ops) == 1:
        op = other_ops[0]
        op_type = op.get("type")
        result = _analyze_table_impl(
            path,
            op_type,
            target_column=op.get("column"),
            sheet_name=sheet_name,
            associated_column=op.get("associated_column"),
        )
        if result.get("status") == "success":
            executed_ids.append(op["operation_id"])
        return result

    # Multi-filter ± sort path
    if other_ops:
        return {
            "status": "plan_error",
            "reason": "unsupported_composed_operation_mix",
        }

    if not filter_ops:
        return {"status": "plan_error", "reason": "no_filter_ops_in_composed_plan"}

    # --- Load table once ---
    file_type = _file_type_from_path(path)
    if file_type in ("unsupported_legacy", "unsupported"):
        return _build_error_result("unsupported", "unsupported_file_type", file_type, path)

    validation = _validate_file_path(path)
    if validation.get("status") == "failure":
        return _build_error_result(
            "blocked", validation.get("reason", "path_safety_blocked"), file_type, path
        )

    full_path = validation["resolved_path"]
    if not os.path.exists(full_path):
        return _build_error_result("not_found", "file_not_found", file_type, path)

    try:
        file_size = os.path.getsize(full_path)
    except OSError:
        return _build_error_result("failed", "file_stat_error", file_type, path)

    if file_size > MAX_FILE_SIZE_BYTES:
        return _build_error_result("unsupported", "analysis_bounds_exceeded", file_type, path)

    if file_type == "csv":
        headers, data_rows, status, reason = _scan_csv(full_path)
        resolved_sheet = None
    else:
        headers, data_rows, status, reason_or_sheet = _scan_xlsx(full_path, sheet_name)
        if status is not None:
            return _build_error_result(status, reason_or_sheet, file_type, path)
        resolved_sheet = reason_or_sheet
        reason = None

    if status is not None:
        return _build_error_result(status, reason, file_type, path, resolved_sheet)
    if not headers:
        return _build_error_result("unsupported", "empty_table", file_type, path, resolved_sheet)

    # --- Apply AND filters sequentially (short-circuit on empty result) ---
    # surviving_indices: set of 0-based indices into data_rows that still match all filters so far.
    surviving_indices: set[int] = set(range(len(data_rows)))
    all_column_refs = []
    combined_warnings: list[str] = []
    combined_limitations: list[str] = []
    predicate_descriptions: list[str] = []

    for op in filter_ops:
        op_id = op["operation_id"]
        col = op.get("column", "")
        f_op = (op.get("filter_op") or "").strip().lower()
        f_val = (op.get("filter_value") or "").strip()
        f_val_to = (op.get("filter_value_to") or "").strip()

        # Use _filter_matching_indices (no truncation) so intermediate AND
        # passes see all surviving rows, not just the display-capped first 1000.
        surviving_list = sorted(surviving_indices)
        current_rows_slice = [data_rows[i] for i in surviving_list]

        idx_result = _filter_matching_indices(
            headers,
            current_rows_slice,
            col,
            f_op,
            f_val,
            f_val_to,
            file_type,
        )
        if idx_result.get("status") != "success":
            return _build_error_result(
                idx_result["status"],
                idx_result["status_reason"],
                file_type,
                path,
                resolved_sheet,
            )

        executed_ids.append(op_id)

        # matching_indices are 0-based positions within current_rows_slice.
        # Map them back to original data_rows indices via surviving_list.
        surviving_indices = {
            surviving_list[pos]
            for pos in idx_result["matching_indices"]
            if pos < len(surviving_list)
        }

        canonical_col = idx_result.get("canonical_col", col)
        all_column_refs.append(f"column:{canonical_col}")
        predicate_descriptions.append(
            _describe_predicate(canonical_col, f_op, f_val, f_val_to)
        )

    # Final matched rows after all AND filters: serialize with original 1-based row numbers.
    final_matched_rows = []
    for orig_idx in sorted(surviving_indices):
        row_number = orig_idx + 1
        final_matched_rows.append(_serialize_row(headers, data_rows[orig_idx], row_number, file_type))

    matched_count = len(final_matched_rows)
    truncated = matched_count > MAX_FILTER_RESULT_ROWS
    returned_rows = final_matched_rows[:MAX_FILTER_RESULT_ROWS]
    returned_count = len(returned_rows)
    result_complete = not truncated

    if truncated:
        combined_limitations.append(
            f"Only the first {MAX_FILTER_RESULT_ROWS:,} of {matched_count:,} "
            f"matched rows are included in the result."
        )

    combined_pred = " AND ".join(predicate_descriptions)

    # --- Sort (optional, single column) ---
    if sort_ops:
        sort_op = sort_ops[0]
        sort_col = sort_op.get("column", "")
        sort_dir = (sort_op.get("direction") or "asc").lower()
        if sort_dir in ("ascending",):
            sort_dir = "asc"
        if sort_dir in ("descending",):
            sort_dir = "desc"

        sort_result = _apply_sort(returned_rows, headers, sort_col, sort_dir, file_type)
        if sort_result.get("status") != "success":
            return _build_error_result(
                sort_result["status"],
                sort_result["status_reason"],
                file_type,
                path,
                resolved_sheet,
            )
        returned_rows = sort_result["sorted_rows"]
        combined_warnings.extend(sort_result.get("warnings", []))
        executed_ids.append(sort_ops[0]["operation_id"])

    # Build answer_text
    if matched_count == 0:
        answer_text = f"No rows match {combined_pred}."
    elif truncated:
        answer_text = (
            f"{matched_count:,} rows match {combined_pred}. "
            f"Showing the first {returned_count:,} rows."
        )
    else:
        name_col = _resolve_auto_name_like_column(headers)
        names = []
        if name_col and matched_count <= 10:
            name_idx = next(
                (i for i, h in enumerate(headers) if h.strip().lower() == name_col.strip().lower()),
                None,
            )
            if name_idx is not None:
                for row_dict in returned_rows:
                    val = next(
                        (c["value"] for c in row_dict.get("cells", []) if c["column_index"] == name_idx + 1),
                        None,
                    )
                    if val and str(val).strip():
                        names.append(str(val).strip())
        if names:
            answer_text = (
                f"{matched_count} row{'s' if matched_count != 1 else ''} match "
                f"{combined_pred}: {', '.join(names)}."
            )
        else:
            answer_text = (
                f"{matched_count} row{'s' if matched_count != 1 else ''} match {combined_pred}."
            )

    if sort_ops:
        sort_col_label = sort_ops[0].get("column", "")
        sort_dir_label = sort_ops[0].get("direction", "asc")
        answer_text += f" Sorted by {sort_col_label} ({sort_dir_label})."

    final_row_refs = [r["row_ref"] for r in returned_rows]
    final_cell_refs = []
    for r in returned_rows:
        for c in r.get("cells", []):
            if c.get("cell_ref"):
                final_cell_refs.append(c["cell_ref"])

    return {
        "status": "success",
        "status_reason": None,
        "operation": "multi_filter" if not sort_ops else "multi_filter_sort",
        "input_file": path,
        "file_type": file_type,
        "sheet_name": resolved_sheet if file_type == "xlsx" else None,
        "rows_evaluated": len(data_rows),
        "matched_row_count": matched_count,
        "returned_row_count": returned_count,
        "result_complete": result_complete,
        "truncated": truncated,
        "rows": returned_rows,
        "column_refs": list(dict.fromkeys(all_column_refs)),
        "row_refs": final_row_refs,
        "cell_refs": final_cell_refs,
        "warnings": combined_warnings,
        "limitations": combined_limitations,
        "answer_text": answer_text,
    }


def _apply_sort(
    serialized_rows: list,
    headers: list,
    column_name: str,
    direction: str,
    file_type: str,
) -> dict:
    """Sort a list of serialized row dicts by a single column deterministically.

    Rules:
    - Explicit column and direction required.
    - Blanks always sort last regardless of direction.
    - Purely numeric columns sort numerically (Decimal).
    - Text columns sort case-insensitively.
    - Mixed-type columns return unsupported to prevent hidden unsafe ordering.
    - Stable sort (preserves original order for equal keys).
    - No locale-dependent behavior.
    """
    col_idx, canonical_col, reason = _resolve_column(headers, column_name, for_what="sort")
    if reason is not None:
        return {
            "status": "not_found" if reason == "column_not_found" else "ambiguous",
            "status_reason": reason,
        }

    col_pos = col_idx  # 1-based

    values = []
    for row_dict in serialized_rows:
        cell = next((c for c in row_dict.get("cells", []) if c["column_index"] == col_pos), None)
        raw_val = cell["value"] if cell else ""
        values.append(raw_val)

    # Classify column from serialized values
    numeric_count = 0
    text_count = 0
    blank_count = 0
    for v in values:
        if v is None or str(v).strip() == "":
            blank_count += 1
        elif _is_numeric_string(str(v).strip()):
            numeric_count += 1
        else:
            text_count += 1

    if numeric_count > 0 and text_count > 0:
        return {
            "status": "unsupported",
            "status_reason": "sort_mixed_type_column",
        }

    use_numeric = numeric_count > 0 and text_count == 0
    reverse = direction == "desc"
    warnings: list[str] = []

    def sort_key(row_dict):
        cell = next(
            (c for c in row_dict.get("cells", []) if c["column_index"] == col_pos), None
        )
        raw = cell["value"] if cell else ""
        is_blank = raw is None or str(raw).strip() == ""
        if is_blank:
            # Blanks always last — use a sentinel that sorts after any value
            return (1, Decimal(0) if use_numeric else "")
        if use_numeric:
            try:
                return (0, Decimal(str(raw).strip()))
            except Exception:
                return (0, Decimal(0))
        return (0, str(raw).casefold())

    sorted_rows = sorted(serialized_rows, key=sort_key, reverse=reverse)
    # Fix: blanks always last regardless of direction — re-sort blanks to end
    non_blank = [r for r in sorted_rows if not _is_blank_sort_row(r, col_pos)]
    blank_rows = [r for r in sorted_rows if _is_blank_sort_row(r, col_pos)]
    sorted_rows = non_blank + blank_rows

    return {
        "status": "success",
        "sorted_rows": sorted_rows,
        "sort_column": canonical_col,
        "sort_direction": direction,
        "warnings": warnings,
    }


def _is_blank_sort_row(row_dict: dict, col_pos: int) -> bool:
    cell = next((c for c in row_dict.get("cells", []) if c["column_index"] == col_pos), None)
    if not cell:
        return True
    v = cell.get("value", "")
    return v is None or str(v).strip() == ""
